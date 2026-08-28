"""
Admin Analytics & Multi-Format Export Service (Stage 3)
======================================================
Reuses ReportLab (PDF), openpyxl (Excel), and Python's CSV engine
to provide enterprise platform reporting and statistical analytics.
"""
import io
import csv
from datetime import datetime, date, timedelta
from decimal import Decimal

from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from django.contrib.auth.models import User
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import inch

from core.models import UserProfile, ActivityLog
from .models import (
    ClientProfile,
    FreelancerProfile,
    MarketplaceProject,
    ProjectApplication,
    ProjectPaymentRecord,
    ProjectReport,
    FreelancerReport,
    MarketplaceDispute,
    PlatformSupportTicket,
    FreelancerVerification,
)


class AdminVerificationService:
    """
    Administrative controls for reviewing and approving freelancer verifications.
    """
    @classmethod
    def approve_verification(cls, verification, admin_user=None, notes=None):
        verification.calculate_profile_completion()
        verification.admin_review_status = 'approved'
        verification.admin_reviewed_at = timezone.now()
        verification.admin_reviewed_by = admin_user
        if notes:
            verification.admin_review_notes = notes
        verification.update_verification_status()
        verification.save()
        return verification

    @classmethod
    def reject_verification(cls, verification, admin_user=None, notes=None):
        verification.admin_review_status = 'rejected'
        verification.admin_reviewed_at = timezone.now()
        verification.admin_reviewed_by = admin_user
        if notes:
            verification.admin_review_notes = notes
        verification.update_verification_status()
        verification.save()
        return verification

    @classmethod
    def suspend_verification(cls, verification, admin_user=None, notes=None):
        verification.admin_review_status = 'suspended'
        verification.admin_reviewed_at = timezone.now()
        verification.admin_reviewed_by = admin_user
        if notes:
            verification.admin_review_notes = notes
        verification.update_verification_status()
        verification.save()
        return verification



class AdminAnalyticsService:
    """
    Computes comprehensive platform statistics, trends, conversion ratios,
    and risk indicators for marketplace administrators.
    """

    @classmethod
    def get_platform_kpis(cls):
        now = timezone.now()
        today = now.date()

        total_clients = ClientProfile.objects.count()
        total_freelancers = FreelancerProfile.objects.count()
        total_users = User.objects.count()

        total_projects = MarketplaceProject.objects.count()
        open_projects = MarketplaceProject.objects.filter(status='open').count()
        active_projects = MarketplaceProject.objects.filter(status__in=['assigned', 'in_progress']).count()
        completed_projects = MarketplaceProject.objects.filter(status='completed').count()
        closed_projects = MarketplaceProject.objects.filter(status='closed').count()

        total_applications = ProjectApplication.objects.count()
        pending_applications = ProjectApplication.objects.filter(status='pending').count()
        accepted_applications = ProjectApplication.objects.filter(status='accepted').count()

        open_client_reports = ProjectReport.objects.filter(status__in=['open', 'under_review']).count()
        open_freelancer_reports = FreelancerReport.objects.filter(status__in=['open', 'under_review']).count()
        total_open_reports = open_client_reports + open_freelancer_reports

        open_disputes = MarketplaceDispute.objects.filter(status__in=['open', 'under_investigation']).count()
        total_disputes = MarketplaceDispute.objects.count()

        suspended_users = UserProfile.objects.filter(is_suspended=True).count()
        reported_users_count = (
            ProjectReport.objects.filter(reported_user__isnull=False).values('reported_user').distinct().count() +
            FreelancerReport.objects.filter(reported_client__isnull=False).values('reported_client').distinct().count()
        )

        open_support_tickets = PlatformSupportTicket.objects.filter(status__in=['open', 'in_progress']).count()

        total_budget_sum = MarketplaceProject.objects.aggregate(total=Sum('budget'))['total'] or Decimal('0.00')
        total_paid_sum = ProjectPaymentRecord.objects.filter(status='paid').aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        total_pending_pay = ProjectPaymentRecord.objects.filter(status__in=['pending', 'partially_paid']).aggregate(
            total=Sum('total_budget') - Sum('amount_paid')
        )['total'] or Decimal('0.00')

        # Conversion rate: Accepted / Total applications
        hire_conversion_rate = (accepted_applications / total_applications * 100) if total_applications > 0 else 0.0

        # Project completion rate
        finished_projects = completed_projects + closed_projects
        completion_rate = (completed_projects / finished_projects * 100) if finished_projects > 0 else (100.0 if completed_projects > 0 else 0.0)

        return {
            'total_clients': total_clients,
            'total_freelancers': total_freelancers,
            'total_users': total_users,
            'total_projects': total_projects,
            'open_projects': open_projects,
            'active_projects': active_projects,
            'completed_projects': completed_projects,
            'closed_projects': closed_projects,
            'total_applications': total_applications,
            'pending_applications': pending_applications,
            'accepted_applications': accepted_applications,
            'open_reports': total_open_reports,
            'open_disputes': open_disputes,
            'total_disputes': total_disputes,
            'suspended_users': suspended_users,
            'reported_users': reported_users_count,
            'open_support_tickets': open_support_tickets,
            'total_budget_volume': float(total_budget_sum),
            'total_paid_volume': float(total_paid_sum),
            'total_pending_payment': float(total_pending_pay),
            'hire_conversion_rate': round(hire_conversion_rate, 1),
            'completion_rate': round(completion_rate, 1),
        }

    @classmethod
    def get_monthly_trends(cls):
        today = timezone.now().date()
        months_labels = []
        project_counts = []
        application_counts = []
        volume_data = []

        for i in range(11, -1, -1):
            m_date = today - timedelta(days=i * 30)
            m_start = date(m_date.year, m_date.month, 1)
            if m_date.month == 12:
                m_end = date(m_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                m_end = date(m_date.year, m_date.month + 1, 1) - timedelta(days=1)

            month_label = m_start.strftime('%b %Y')
            months_labels.append(month_label)

            p_cnt = MarketplaceProject.objects.filter(created_at__date__range=[m_start, m_end]).count()
            project_counts.append(p_cnt)

            a_cnt = ProjectApplication.objects.filter(created_at__date__range=[m_start, m_end]).count()
            application_counts.append(a_cnt)

            vol = MarketplaceProject.objects.filter(created_at__date__range=[m_start, m_end]).aggregate(total=Sum('budget'))['total'] or 0
            volume_data.append(float(vol))

        return {
            'labels': months_labels,
            'projects': project_counts,
            'applications': application_counts,
            'volume': volume_data,
        }

    @classmethod
    def get_predictive_risk_analysis(cls):
        """
        Calculates project delay risk with small-dataset safety guard.
        Requires at least 5 projects to produce statistical estimates.
        """
        total_projects = MarketplaceProject.objects.count()
        min_required = 5
        is_statistically_significant = total_projects >= min_required

        overdue_projects = []
        high_risk_projects = []

        today = timezone.now().date()
        for p in MarketplaceProject.objects.filter(status__in=['open', 'assigned', 'in_progress']).select_related('client', 'assigned_freelancer'):
            is_overdue = p.deadline and (today > p.deadline)
            days_left = (p.deadline - today).days if p.deadline else 999
            
            # Risk criteria: overdue or <3 days left with progress < 50%
            risk_score = 0
            if is_overdue:
                risk_score = 90
            elif days_left <= 3 and p.progress < 50:
                risk_score = 75
            elif days_left <= 7 and p.progress < 30:
                risk_score = 50

            item = {
                'id': str(p.id),
                'title': p.title,
                'client': p.client.display_name,
                'freelancer': p.assigned_freelancer.username if p.assigned_freelancer else 'Unassigned',
                'budget': float(p.budget or 0),
                'deadline': p.deadline.strftime('%Y-%m-%d') if p.deadline else 'None',
                'progress': p.progress,
                'risk_score': risk_score,
                'is_overdue': is_overdue,
            }
            if is_overdue:
                overdue_projects.append(item)
            if risk_score >= 50:
                high_risk_projects.append(item)

        return {
            'is_ready': is_statistically_significant,
            'min_records_required': min_required,
            'total_analyzed': total_projects,
            'overdue_count': len(overdue_projects),
            'high_risk_count': len(high_risk_projects),
            'high_risk_projects': high_risk_projects[:10],
        }


class AdminExportService:
    """
    Generates presentation-ready PDF, Excel, and CSV downloads
    for platform audit, client, freelancer, project, and dispute reports.
    """

    @classmethod
    def export_csv(cls, filename, headers, rows):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        return response

    @classmethod
    def export_excel(cls, filename, sheet_title, headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]

        # Header styling
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A148C', end_color='4A148C', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        border_thin = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        ws.append(headers)
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for row_data in rows:
            ws.append(row_data)

        # Apply borders and auto-fit width
        for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border_thin

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    @classmethod
    def export_pdf(cls, filename, title, headers, rows):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=0.4 * inch,
            rightMargin=0.4 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'AdminDocTitle',
            parent=styles['Heading1'],
            fontSize=16,
            leading=20,
            textColor=colors.HexColor('#2E1065'),
            alignment=TA_CENTER
        )
        meta_style = ParagraphStyle(
            'AdminDocMeta',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#64748B'),
            alignment=TA_CENTER
        )
        cell_style = ParagraphStyle(
            'AdminDocCell',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#1E293B')
        )
        th_style = ParagraphStyle(
            'AdminDocTh',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            fontName='Helvetica-Bold',
            textColor=colors.white,
            alignment=TA_CENTER
        )

        elements = [
            Paragraph(f"<b>FREELANCER MARKETPLACE INTELLIGENCE</b>", title_style),
            Spacer(1, 4),
            Paragraph(f"Executive Report: {title} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M UTC')}", meta_style),
            Spacer(1, 10),
            HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#7C3AED'), spaceAfter=12),
        ]

        table_data = [[Paragraph(h, th_style) for h in headers]]
        for row in rows:
            table_data.append([Paragraph(str(c if c is not None else ''), cell_style) for c in row])

        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4C1D95')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        elements.append(t)
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
