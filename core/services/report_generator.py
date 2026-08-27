"""
Freelancer Intelligence Platform — Enterprise Report Generator Service
========================================================================
Generates presentation-ready executive PDF reports (ReportLab), formatted
multi-sheet Excel workbooks (openpyxl), and authorized CSV data streams
incorporating full branding, filter context, and business intelligence.
"""

import io
from datetime import datetime, date, timedelta
from decimal import Decimal

from django.utils import timezone
from django.db.models import Sum, Avg, Count, Min, Max, Q, F

from core.models import Client, Project, Payment, Task, Income, Expense, Invoice
from core.services.analytics_engine import DataAnalyticsEngine
from core.services.data_science_service import DataScienceService


def _to_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


class ReportGeneratorService:
    """
    Service providing comprehensive multi-format report exports for
    Freelancer Intelligence Platform.
    """

    def __init__(self, user, filters=None):
        self.user = user
        self.filters = filters or {}
        self.analytics_engine = DataAnalyticsEngine(user, filters=self.filters)
        self.ds_service = DataScienceService(user)

    # --------------------------------------------------------------------------
    # 1. COMPREHENSIVE PDF REPORT (ReportLab)
    # --------------------------------------------------------------------------
    def generate_comprehensive_pdf(self):
        """
        Generates an executive-ready multi-page PDF report.
        """
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.units import inch

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        styles = getSampleStyleSheet()
        primary_color = colors.HexColor('#DB9941')
        dark_bg = colors.HexColor('#07111D')
        text_dark = colors.HexColor('#1E293B')
        muted_gray = colors.HexColor('#64748B')
        border_color = colors.HexColor('#CBD5E1')

        # Custom Styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            textColor=primary_color,
            fontName='Helvetica-Bold',
            spaceAfter=2
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            textColor=muted_gray,
            fontName='Helvetica',
            spaceAfter=12
        )
        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=13,
            leading=16,
            textColor=dark_bg,
            fontName='Helvetica-Bold',
            spaceBefore=10,
            spaceAfter=6
        )
        body_style = ParagraphStyle(
            'ReportBody',
            parent=styles['Normal'],
            fontSize=8.5,
            leading=12,
            textColor=text_dark,
            fontName='Helvetica'
        )
        bold_body = ParagraphStyle(
            'BoldReportBody',
            parent=body_style,
            fontName='Helvetica-Bold'
        )

        elements = []
        today_str = timezone.now().strftime('%B %d, %Y')

        # ── Header Banner ──
        header_data = [
            [
                Paragraph("<b>FREELANCER INTELLIGENCE PLATFORM</b><br/><font size='8' color='#64748B'>Manage. Analyze. Predict. Grow.</font>", body_style),
                Paragraph(f"<b>Executive BI Report</b><br/><font size='8' color='#64748B'>Generated: {today_str}</font>", ParagraphStyle('RightH', parent=body_style, alignment=TA_RIGHT))
            ]
        ]
        header_table = Table(header_data, colWidths=[3.5 * inch, 3.5 * inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1.5, color=primary_color, spaceBefore=4, spaceAfter=10))

        # ── Filter Context Block ──
        filter_items = []
        if self.filters.get('client'):
            filter_items.append(f"Client: {self.filters['client']}")
        if self.filters.get('status'):
            filter_items.append(f"Status: {self.filters['status'].capitalize()}")
        if self.filters.get('start_date') or self.filters.get('end_date'):
            filter_items.append(f"Date Range: {self.filters.get('start_date', 'Start')} to {self.filters.get('end_date', 'Present')}")
        
        filter_str = " | ".join(filter_items) if filter_items else "All Time Portfolio (No constraints applied)"

        filter_data = [[
            Paragraph(f"<b>FILTERS APPLIED:</b> {filter_str}", ParagraphStyle('Filt', parent=body_style, textColor=dark_bg, fontSize=8))
        ]]
        filter_table = Table(filter_data, colWidths=[7.0 * inch])
        filter_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(filter_table)
        elements.append(Spacer(1, 8))

        # ── Section 1: Executive KPI Summary ──
        elements.append(Paragraph("1. Executive Business Performance Summary", section_heading))
        
        projects = self.analytics_engine.projects_qs
        payments = self.analytics_engine.payments_qs
        
        total_p = projects.count()
        completed_p = projects.filter(status='completed').count()
        comp_rate = round((completed_p / total_p * 100), 1) if total_p > 0 else 0.0

        total_rev = _to_float(payments.filter(status='paid').aggregate(t=Sum('amount'))['t'] or 0.0)
        pending_rev = _to_float(payments.filter(status='pending').aggregate(t=Sum('amount'))['t'] or 0.0)
        avg_budget = _to_float(projects.aggregate(a=Avg('budget'))['a'] or 0.0)

        kpi_table_data = [
            ["Total Realized Revenue", "Pending Invoices", "Avg. Project Value", "Completion Rate", "Total Projects"],
            [f"${total_rev:,.0f}", f"${pending_rev:,.0f}", f"${avg_budget:,.0f}", f"{comp_rate}%", str(total_p)]
        ]
        kpi_table = Table(kpi_table_data, colWidths=[1.4 * inch] * 5)
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), dark_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#16A34A')),
            ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#D97706')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 10))

        # ── Section 2: Data Quality & Statistical Profile ──
        elements.append(Paragraph("2. Data Quality & Statistical Health", section_heading))
        quality = self.analytics_engine.get_data_quality_audit()
        eda = self.analytics_engine.get_exploratory_analysis()
        b_stats = eda.get('budget_stats', {})

        quality_data = [
            ["Data Quality Score", "Integrity Status", "Budget Mean", "Budget Median (Q2)", "Budget Std. Dev"],
            [
                f"{quality['score']}/100",
                quality['grade'],
                f"${b_stats.get('mean', 0):,.0f}" if b_stats.get('mean') else "N/A",
                f"${b_stats.get('median', 0):,.0f}" if b_stats.get('median') else "N/A",
                f"${b_stats.get('std_dev', 0):,.0f}" if b_stats.get('std_dev') else "N/A"
            ]
        ]
        quality_table = Table(quality_data, colWidths=[1.4 * inch] * 5)
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(quality_table)
        elements.append(Spacer(1, 10))

        # ── Section 3: Time-Series Revenue Forecasting ──
        elements.append(Paragraph("3. Predictive Time-Series Forecast (Estimated)", section_heading))
        forecast = self.ds_service.get_time_series_forecast(forecast_months=3)

        if forecast.get('is_sufficient'):
            fc_rows = [["Horizon Month", "Algorithm", "Estimated Forecast ($)", "95% Confidence Interval"]]
            for i, m in enumerate(forecast.get('forecast_labels', [])):
                val = forecast['forecast_values'][i]
                low = forecast['lower_bounds'][i]
                high = forecast['upper_bounds'][i]
                fc_rows.append([m, "Holt-Winters + OLS Ensemble", f"${val:,.0f}", f"[${low:,.0f} — ${high:,.0f}]"])

            fc_table = Table(fc_rows, colWidths=[1.5 * inch, 2.3 * inch, 1.6 * inch, 1.6 * inch])
            fc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), primary_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), dark_bg),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, border_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(fc_table)
        else:
            elements.append(Paragraph("<i>Time-series forecasting accumulating baseline observations (Requires N ≥ 3 monthly records).</i>", body_style))

        elements.append(Spacer(1, 10))

        # ── Section 4: Key Insights & Recommendations ──
        elements.append(Paragraph("4. Automated Key Insights & Strategic Actions", section_heading))
        insights = self.analytics_engine.generate_automated_insights()
        
        for ins in insights.get('insights', [])[:3]:
            elements.append(Paragraph(f"• <b>{ins['title']}:</b> {ins['summary']} <font color='#64748B' size='7'>[{ins['metric_value']}]</font>", body_style))
            elements.append(Spacer(1, 2))

        elements.append(Spacer(1, 4))
        for rec in insights.get('recommendations', [])[:3]:
            elements.append(Paragraph(f"&bull; <font color='#D97706'><b>ACTION ({rec['priority'].upper()}):</b></font> {rec['title']} &mdash; {rec['detail']}", body_style))
            elements.append(Spacer(1, 2))

        # ── Footer Signoff ──
        elements.append(Spacer(1, 12))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=border_color, spaceBefore=4, spaceAfter=6))
        elements.append(Paragraph(
            "<b>CONFIDENTIAL &bull; FREELANCER INTELLIGENCE PLATFORM</b> &bull; Real Database Verification &bull; All statistical forecasts represent probabilistic estimates.",
            ParagraphStyle('Foot', parent=body_style, fontSize=7, textColor=muted_gray, alignment=TA_CENTER)
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    # --------------------------------------------------------------------------
    # 2. MULTI-SHEET EXCEL EXPORT (openpyxl)
    # --------------------------------------------------------------------------
    def generate_comprehensive_excel(self):
        """
        Generates a formatted multi-sheet Excel workbook using openpyxl.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="07111D", end_color="07111D", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        title_font = Font(name="Arial", size=14, bold=True, color="DB9941")
        bold_font = Font(name="Arial", size=10, bold=True)
        normal_font = Font(name="Arial", size=9)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # ── Sheet 1: Executive Summary ──
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        ws_exec.views.sheetView[0].showGridLines = True

        ws_exec.append(["FREELANCER INTELLIGENCE PLATFORM — EXECUTIVE SUMMARY"])
        ws_exec.cell(1, 1).font = title_font
        ws_exec.append([f"Report Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')} | User: {self.user.username}"])
        ws_exec.cell(2, 1).font = normal_font
        ws_exec.append([])

        projects = self.analytics_engine.projects_qs
        payments = self.analytics_engine.payments_qs

        total_rev = _to_float(payments.filter(status='paid').aggregate(t=Sum('amount'))['t'] or 0.0)
        pending_rev = _to_float(payments.filter(status='pending').aggregate(t=Sum('amount'))['t'] or 0.0)
        avg_budget = _to_float(projects.aggregate(a=Avg('budget'))['a'] or 0.0)
        total_p = projects.count()
        comp_p = projects.filter(status='completed').count()

        kpi_rows = [
            ["Metric", "Value", "Description"],
            ["Total Realized Revenue", f"${total_rev:,.2f}", "Total collected payments"],
            ["Accounts Receivable", f"${pending_rev:,.2f}", "Unpaid & pending invoices"],
            ["Average Project Value", f"${avg_budget:,.2f}", "Mean contracted project budget"],
            ["Total Tracked Projects", total_p, "Total active and completed projects"],
            ["Completed Projects", comp_p, "Successfully delivered projects"],
            ["Completion Velocity", f"{(comp_p / total_p * 100):.1f}%" if total_p > 0 else "0%", "Portfolio completion percentage"]
        ]

        for r in kpi_rows:
            ws_exec.append(r)

        # ── Sheet 2: Projects ──
        ws_proj = wb.create_sheet(title="Projects")
        ws_proj.views.sheetView[0].showGridLines = True
        ws_proj.append(["Project Name", "Client", "Status", "Priority", "Progress (%)", "Budget ($)", "Deadline", "Created Date"])
        
        for p in projects.order_by('-created_at'):
            ws_proj.append([
                p.name,
                p.client.name if p.client else "N/A",
                p.get_status_display(),
                p.get_priority_display(),
                p.progress,
                _to_float(p.budget),
                str(p.deadline or "N/A"),
                str(p.created_at.date())
            ])

        # ── Sheet 3: Payments ──
        ws_pay = wb.create_sheet(title="Payments")
        ws_pay.views.sheetView[0].showGridLines = True
        ws_pay.append(["Invoice #", "Project", "Client", "Amount ($)", "Status", "Payment Method", "Due Date", "Paid Date"])
        
        for pm in payments.order_by('-created_at'):
            ws_pay.append([
                pm.invoice_number or "N/A",
                pm.project.name if pm.project else "N/A",
                pm.project.client.name if (pm.project and pm.project.client) else "N/A",
                _to_float(pm.amount),
                pm.get_status_display(),
                pm.get_payment_method_display() if hasattr(pm, 'get_payment_method_display') else pm.payment_method,
                str(pm.due_date or "N/A"),
                str(pm.paid_date or "N/A")
            ])

        # ── Sheet 4: Forecast & Insights ──
        ws_fc = wb.create_sheet(title="Forecast & Insights")
        ws_fc.views.sheetView[0].showGridLines = True
        ws_fc.append(["Horizon Month", "Forecast Algorithm", "Estimated Forecast ($)", "95% Lower Bound", "95% Upper Bound"])

        forecast = self.ds_service.get_time_series_forecast(forecast_months=6)
        if forecast.get('is_sufficient'):
            for i, m in enumerate(forecast['forecast_labels']):
                ws_fc.append([
                    m,
                    forecast['method_used'],
                    forecast['forecast_values'][i],
                    forecast['lower_bounds'][i],
                    forecast['upper_bounds'][i]
                ])

        # Format Header Cells and Auto-fit column widths across all sheets
        for sheet in wb.worksheets:
            first_row = sheet[1] if sheet.title != "Executive Summary" else sheet[4]
            for cell in first_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # --------------------------------------------------------------------------
    # 3. CSV DATA EXPORT
    # --------------------------------------------------------------------------
    def generate_csv_stream(self, dataset_type):
        """
        Streams clean CSV output for projects, payments, or client datasets.
        """
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        if dataset_type == 'projects':
            writer.writerow(['ID', 'Name', 'Client', 'Status', 'Priority', 'Progress', 'Budget', 'Deadline', 'Created At'])
            for p in self.analytics_engine.projects_qs:
                writer.writerow([
                    str(p.id),
                    p.name,
                    p.client.name if p.client else '',
                    p.status,
                    p.priority,
                    p.progress,
                    _to_float(p.budget),
                    str(p.deadline or ''),
                    str(p.created_at)
                ])

        elif dataset_type == 'payments':
            writer.writerow(['ID', 'Invoice Number', 'Project', 'Amount', 'Status', 'Method', 'Due Date', 'Paid Date'])
            for pm in self.analytics_engine.payments_qs:
                writer.writerow([
                    str(pm.id),
                    pm.invoice_number or '',
                    pm.project.name if pm.project else '',
                    _to_float(pm.amount),
                    pm.status,
                    pm.payment_method,
                    str(pm.due_date or ''),
                    str(pm.paid_date or '')
                ])

        elif dataset_type == 'clients':
            writer.writerow(['ID', 'Name', 'Email', 'Company', 'Status', 'Phone', 'Created At'])
            for c in self.analytics_engine.clients_qs:
                writer.writerow([
                    str(c.id),
                    c.name,
                    c.email,
                    c.company or '',
                    c.status,
                    c.phone or '',
                    str(c.created_at)
                ])

        buffer.seek(0)
        return buffer
