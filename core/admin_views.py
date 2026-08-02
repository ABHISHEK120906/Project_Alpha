import os
import json
import csv
import io
from datetime import datetime, date, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.db import models, connection
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse, JsonResponse, FileResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.serializers import serialize
from django.contrib.sessions.models import Session
from django.core.management import call_command

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import (
    Client, Project, Payment, Task, Note, ActivityLog, UserProfile,
    ProjectFile, ProjectComment, Income, Expense, Invoice, InvoiceItem,
    CalendarEvent, Notification, LoginHistory, BlockedIP, SystemSetting,
    RefundRequest, Announcement
)


# ============================================================
# AUTHORIZATION DECORATOR
# ============================================================

def admin_required(view_func):
    """
    Decorator enforcing Super Admin / Staff authorization.
    Rejects non-admin users with HTTP 403 Forbidden or redirect.
    """
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f"/login/?next={request.path}")
        
        profile = getattr(request.user, 'profile', None)
        is_admin = (request.user.is_staff or request.user.is_superuser or (profile and profile.role == 'admin'))
        
        if not is_admin:
            messages.error(request, "403 Forbidden: Super Admin access required.")
            return redirect('core:forbidden')
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def log_admin_activity(user, action, model_type, model_id, description, request=None):
    """Helper to log administrative actions to ActivityLog."""
    ip_address = None
    user_agent = None
    if request:
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        ip_address = x_forwarded_for.split(',')[0].strip() if x_forwarded_for else request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:255]

    ActivityLog.objects.create(
        user=user,
        action=action,
        model_type=model_type,
        model_id=model_id,
        description=f"[ADMIN] {description}",
        ip_address=ip_address,
        user_agent=user_agent
    )


# ============================================================
# FEATURE 3 & 14 — SUPER ADMIN DASHBOARD
# ============================================================

@admin_required
def admin_dashboard(request):
    """
    Super Admin Master Dashboard:
    Computes 21 KPI Stat Cards & 11 Visual Analytics Datasets.
    """
    now = timezone.now()
    today = now.date()
    start_of_month = date(today.year, today.month, 1)

    # ── 21 STAT KPI CARDS ─────────────────────────────────────
    total_users = User.objects.count()
    
    # Online users: logged in within past 15 minutes
    fifteen_mins_ago = now - timedelta(minutes=15)
    online_users = UserProfile.objects.filter(last_login_at__gte=fifteen_mins_ago).count()
    offline_users = max(0, total_users - online_users)
    active_users = UserProfile.objects.filter(is_suspended=False, is_deleted=False, user__is_active=True).count()
    blocked_users = UserProfile.objects.filter(Q(is_suspended=True) | Q(user__is_active=False)).count()
    verified_users = UserProfile.objects.filter(is_verified=True).count()

    total_projects = Project.objects.filter(is_archived=False).count()
    running_projects = Project.objects.filter(status='in_progress', is_archived=False).count()
    pending_projects = Project.objects.filter(status='pending', is_archived=False).count()
    completed_projects = Project.objects.filter(status='completed', is_archived=False).count()
    cancelled_projects = Project.objects.filter(status='cancelled', is_archived=False).count()

    total_clients = Client.objects.filter(is_archived=False).count()

    # Revenue calculations (Payments status='paid' + Incomes)
    paid_payments_total = Payment.objects.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    incomes_total = Income.objects.aggregate(total=Sum('amount'))['total'] or 0
    total_revenue = float(paid_payments_total + incomes_total)

    monthly_payments = Payment.objects.filter(status='paid', paid_date__gte=start_of_month).aggregate(total=Sum('amount'))['total'] or 0
    monthly_incomes = Income.objects.filter(date__gte=start_of_month).aggregate(total=Sum('amount'))['total'] or 0
    monthly_revenue = float(monthly_payments + monthly_incomes)

    total_expenses = float(Expense.objects.aggregate(total=Sum('amount'))['total'] or 0)
    net_profit = total_revenue - total_expenses

    pending_payments_val = float(Payment.objects.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0)
    completed_payments_count = Payment.objects.filter(status='paid').count()

    today_payments = Payment.objects.filter(status='paid', paid_date=today).aggregate(total=Sum('amount'))['total'] or 0
    today_incomes = Income.objects.filter(date=today).aggregate(total=Sum('amount'))['total'] or 0
    today_revenue = float(today_payments + today_incomes)

    total_files_uploaded = ProjectFile.objects.count()

    kpi_stats = {
        'total_users': total_users,
        'online_users': online_users,
        'offline_users': offline_users,
        'active_users': active_users,
        'blocked_users': blocked_users,
        'verified_users': verified_users,
        'total_projects': total_projects,
        'running_projects': running_projects,
        'pending_projects': pending_projects,
        'completed_projects': completed_projects,
        'cancelled_projects': cancelled_projects,
        'total_clients': total_clients,
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'pending_payments': pending_payments_val,
        'completed_payments': completed_payments_count,
        'today_revenue': today_revenue,
        'this_month_revenue': monthly_revenue,
        'total_files_uploaded': total_files_uploaded,
    }

    # ── 11 VISUAL CHARTS DATA ─────────────────────────────────
    months_labels = []
    revenue_chart_data = []
    expense_chart_data = []
    user_growth_data = []
    registration_trend_data = []

    for i in range(11, -1, -1):
        m_date = today - timedelta(days=i*30)
        m_start = date(m_date.year, m_date.month, 1)
        if m_date.month == 12:
            m_end = date(m_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            m_end = date(m_date.year, m_date.month + 1, 1) - timedelta(days=1)

        month_name = m_start.strftime('%b %Y')
        months_labels.append(month_name)

        # Revenue
        m_pay = Payment.objects.filter(status='paid', paid_date__range=[m_start, m_end]).aggregate(total=Sum('amount'))['total'] or 0
        m_inc = Income.objects.filter(date__range=[m_start, m_end]).aggregate(total=Sum('amount'))['total'] or 0
        revenue_chart_data.append(float(m_pay + m_inc))

        # Expense
        m_exp = Expense.objects.filter(date__range=[m_start, m_end]).aggregate(total=Sum('amount'))['total'] or 0
        expense_chart_data.append(float(m_exp))

        # User growth
        u_count = User.objects.filter(date_joined__lte=timezone.make_aware(datetime.combine(m_end, datetime.max.time()))).count()
        user_growth_data.append(u_count)

        # Registrations in that specific month
        r_count = User.objects.filter(date_joined__date__range=[m_start, m_end]).count()
        registration_trend_data.append(r_count)

    # Project Status & Priority Distributions
    project_status_counts = [
        Project.objects.filter(status='pending', is_archived=False).count(),
        Project.objects.filter(status='in_progress', is_archived=False).count(),
        Project.objects.filter(status='completed', is_archived=False).count(),
        Project.objects.filter(status='on_hold', is_archived=False).count(),
        Project.objects.filter(status='cancelled', is_archived=False).count(),
    ]

    project_priority_counts = [
        Project.objects.filter(priority='low', is_archived=False).count(),
        Project.objects.filter(priority='medium', is_archived=False).count(),
        Project.objects.filter(priority='high', is_archived=False).count(),
        Project.objects.filter(priority='urgent', is_archived=False).count(),
    ]

    # Top Clients & Top Freelancers
    top_clients_qs = Client.objects.annotate(
        total_spent=Sum('projects__payments__amount', filter=Q(projects__payments__status='paid'))
    ).order_by('-total_spent')[:5]
    
    top_clients_names = [c.name for c in top_clients_qs]
    top_clients_spent = [float(c.total_spent or 0) for c in top_clients_qs]

    top_freelancers_qs = User.objects.annotate(
        total_rev=Sum('payments__amount', filter=Q(payments__status='paid'))
    ).order_by('-total_rev')[:5]

    top_freelancers_names = [u.username for u in top_freelancers_qs]
    top_freelancers_rev = [float(u.total_rev or 0) for u in top_freelancers_qs]

    charts_json = {
        'months': months_labels,
        'revenue': revenue_chart_data,
        'expenses': expense_chart_data,
        'project_status': project_status_counts,
        'project_priority': project_priority_counts,
        'user_growth': user_growth_data,
        'registrations': registration_trend_data,
        'top_clients_names': top_clients_names,
        'top_clients_spent': top_clients_spent,
        'top_freelancers_names': top_freelancers_names,
        'top_freelancers_rev': top_freelancers_rev,
    }

    # Recent Feeds & Widgets
    recent_activities = ActivityLog.objects.select_related('user').all()[:8]
    recent_users = User.objects.select_related('profile').order_by('-date_joined')[:5]
    recent_payments = Payment.objects.select_related('user', 'project').order_by('-created_at')[:5]
    recent_projects = Project.objects.select_related('user', 'client').order_by('-created_at')[:5]
    pinned_announcements = Announcement.objects.filter(is_pinned=True)[:3]

    context = {
        'kpi': kpi_stats,
        'charts_json': json.dumps(charts_json),
        'recent_activities': recent_activities,
        'recent_users': recent_users,
        'recent_payments': recent_payments,
        'recent_projects': recent_projects,
        'pinned_announcements': pinned_announcements,
        'page_title': 'Super Admin Master Dashboard',
    }
    return render(request, 'admin_dashboard/dashboard.html', context)


# ============================================================
# FEATURE 4 — COMPLETE USER MANAGEMENT
# ============================================================

@admin_required
def admin_users_list(request):
    """View, search, filter, and manage all users."""
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')
    role_filter = request.GET.get('role', '')

    users_qs = User.objects.select_related('profile').order_by('-date_joined')

    if search_query:
        users_qs = users_qs.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    if role_filter:
        if role_filter == 'admin':
            users_qs = users_qs.filter(Q(is_staff=True) | Q(is_superuser=True) | Q(profile__role='admin'))
        elif role_filter == 'user':
            users_qs = users_qs.filter(is_staff=False, is_superuser=False, profile__role='user')

    if status_filter:
        if status_filter == 'active':
            users_qs = users_qs.filter(is_active=True, profile__is_suspended=False)
        elif status_filter == 'suspended':
            users_qs = users_qs.filter(profile__is_suspended=True)
        elif status_filter == 'deactivated':
            users_qs = users_qs.filter(is_active=False)

    paginator = Paginator(users_qs, 15)
    page_number = request.GET.get('page')
    users = paginator.get_page(page_number)

    context = {
        'users': users,
        'search_query': search_query,
        'status_filter': status_filter,
        'role_filter': role_filter,
        'total_count': users_qs.count(),
        'page_title': 'User Management',
    }
    return render(request, 'admin_dashboard/users/user_list.html', context)


@admin_required
def admin_user_create(request):
    """Create a new user account with role specification."""
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        role = request.POST.get('role', 'user')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()

        if User.objects.filter(username=username).exists():
            messages.error(request, f"Username '{username}' already exists.")
            return redirect('core:admin_users_list')

        if email and User.objects.filter(email__iexact=email).exists():
            messages.error(request, f"Email '{email}' is already registered.")
            return redirect('core:admin_users_list')

        if role == 'admin' and username.lower() != 'svathi':
            messages.error(request, "Action rejected: Only the existing Admin ('Svathi') is allowed to have administrator privileges.")
            return redirect('core:admin_users_list')

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.role = 'user'
        profile.is_verified = True
        profile.save()

        log_admin_activity(request.user, 'create', 'user', user.id, f"Created user {username} (user)", request)
        messages.success(request, f"User '{username}' created successfully!")
        return redirect('core:admin_users_list')

    return redirect('core:admin_users_list')


@admin_required
def admin_user_edit(request, user_id):
    """Edit user details & profile."""
    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)

    if request.method == 'POST':
        user.email = request.POST.get('email', user.email).strip()
        user.first_name = request.POST.get('first_name', user.first_name).strip()
        user.last_name = request.POST.get('last_name', user.last_name).strip()

        new_role = request.POST.get('role', profile.role)
        if new_role == 'admin' and user.username.lower() != 'svathi':
            messages.error(request, "Action rejected: Cannot promote user to Admin. Only 'Svathi' can be Admin.")
            return redirect('core:admin_users_list')

        if user.username.lower() == 'svathi':
            profile.role = 'admin'
            user.is_staff = True
            user.is_superuser = True
        else:
            profile.role = 'user'
            user.is_staff = False
            user.is_superuser = False

        user.save()
        profile.save()

        log_admin_activity(request.user, 'update', 'user', user.id, f"Updated user {user.username}", request)
        messages.success(request, f"User '{user.username}' updated successfully.")
        return redirect('core:admin_users_list')

    return redirect('core:admin_users_list')


@admin_required
def admin_user_suspend(request, user_id):
    """Toggle suspend/ban status for user."""
    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)

    profile.is_suspended = not profile.is_suspended
    profile.save()

    status_str = "suspended/banned" if profile.is_suspended else "unbanned/restored"
    log_admin_activity(request.user, 'status_change', 'user', user.id, f"{status_str.title()} user {user.username}", request)
    messages.success(request, f"User '{user.username}' has been {status_str}.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_activate(request, user_id):
    """Toggle is_active state for user."""
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()

    status_str = "activated" if user.is_active else "deactivated"
    log_admin_activity(request.user, 'status_change', 'user', user.id, f"{status_str.title()} user {user.username}", request)
    messages.success(request, f"User '{user.username}' has been {status_str}.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_verify_email(request, user_id):
    """Toggle email verification status for user."""
    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.is_verified = not profile.is_verified
    profile.save()

    status_str = "verified" if profile.is_verified else "unverified"
    messages.success(request, f"Email verification for '{user.username}' set to {status_str}.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_reset_password(request, user_id):
    """Admin forced password reset for user."""
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '')
        if len(new_password) < 6:
            messages.error(request, "Password must be at least 6 characters.")
            return redirect('core:admin_users_list')

        user.set_password(new_password)
        user.save()
        log_admin_activity(request.user, 'update', 'user', user.id, f"Reset password for user {user.username}", request)
        messages.success(request, f"Password for '{user.username}' reset successfully.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_force_logout(request, user_id):
    """Force logout user by flushing active sessions."""
    user = get_object_or_404(User, id=user_id)
    sessions = Session.objects.all()
    count = 0
    for s in sessions:
        data = s.get_decoded()
        if str(data.get('_auth_user_id')) == str(user.id):
            s.delete()
            count += 1
    log_admin_activity(request.user, 'logout', 'user', user.id, f"Force logged out user {user.username} ({count} sessions terminated)", request)
    messages.success(request, f"Terminated {count} active session(s) for user '{user.username}'.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_delete(request, user_id):
    """Soft or hard delete user."""
    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)
    
    soft_delete = request.GET.get('hard', '0') != '1'
    if soft_delete:
        profile.is_deleted = True
        user.is_active = False
        user.save()
        profile.save()
        log_admin_activity(request.user, 'delete', 'user', user.id, f"Soft-deleted user {user.username}", request)
        messages.success(request, f"User '{user.username}' soft-deleted (account deactivated).")
    else:
        username = user.username
        user.delete()
        log_admin_activity(request.user, 'delete', 'user', user_id, f"Hard-deleted user {username}", request)
        messages.success(request, f"User '{username}' permanently deleted.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_restore(request, user_id):
    """Restore soft-deleted user."""
    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.is_deleted = False
    profile.is_suspended = False
    user.is_active = True
    user.save()
    profile.save()

    log_admin_activity(request.user, 'restore', 'user', user.id, f"Restored deleted user {user.username}", request)
    messages.success(request, f"User '{user.username}' restored successfully.")
    return redirect('core:admin_users_list')


@admin_required
def admin_user_history(request, user_id):
    """View detailed history breakdown for user (AJAX JSON)."""
    user = get_object_or_404(User, id=user_id)
    logins = LoginHistory.objects.filter(user=user)[:15]
    activities = ActivityLog.objects.filter(user=user)[:15]
    projects = Project.objects.filter(user=user)[:10]
    payments = Payment.objects.filter(user=user)[:10]
    files = ProjectFile.objects.filter(user=user)[:10]

    data = {
        'username': user.username,
        'email': user.email,
        'logins': [{'ip': l.ip_address, 'status': l.status, 'time': l.timestamp.strftime('%Y-%m-%d %H:%M'), 'device': l.device} for l in logins],
        'activities': [{'action': a.action, 'description': a.description, 'time': a.timestamp.strftime('%Y-%m-%d %H:%M')} for a in activities],
        'projects_count': projects.count(),
        'payments_count': payments.count(),
        'files_count': files.count(),
    }
    return JsonResponse(data)


# ============================================================
# FEATURE 5 — PROJECT MANAGEMENT (ADMIN SCOPE)
# ============================================================

@admin_required
def admin_projects_list(request):
    """Global Project Catalog for Super Admin."""
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')

    projects_qs = Project.objects.select_related('user', 'client').order_by('-created_at')

    if search_query:
        projects_qs = projects_qs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(client__name__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    if status_filter:
        projects_qs = projects_qs.filter(status=status_filter)
    if priority_filter:
        projects_qs = projects_qs.filter(priority=priority_filter)

    paginator = Paginator(projects_qs, 15)
    page = request.GET.get('page')
    projects = paginator.get_page(page)

    all_users = User.objects.filter(is_active=True)
    all_clients = Client.objects.filter(is_archived=False)

    context = {
        'projects': projects,
        'all_users': all_users,
        'all_clients': all_clients,
        'search_query': search_query,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'page_title': 'Global Project Management',
    }
    return render(request, 'admin_dashboard/projects/project_list.html', context)


@admin_required
def admin_project_create(request):
    """Admin create project assigned to any user and client."""
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        client_id = request.POST.get('client_id')
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '')
        status = request.POST.get('status', 'pending')
        priority = request.POST.get('priority', 'medium')
        budget = request.POST.get('budget', 0)
        start_date = request.POST.get('start_date') or None
        deadline = request.POST.get('deadline') or None

        user = get_object_or_404(User, id=user_id)
        client = get_object_or_404(Client, id=client_id)

        project = Project.objects.create(
            user=user,
            client=client,
            name=name,
            description=description,
            status=status,
            priority=priority,
            budget=budget,
            start_date=start_date,
            deadline=deadline
        )
        log_admin_activity(request.user, 'create', 'project', project.id, f"Created project '{name}' for user {user.username}", request)
        messages.success(request, f"Project '{name}' created successfully.")
    return redirect('core:admin_projects_list')


@admin_required
def admin_project_edit(request, project_id):
    """Admin edit project parameters."""
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        project.name = request.POST.get('name', project.name).strip()
        project.description = request.POST.get('description', project.description)
        project.status = request.POST.get('status', project.status)
        project.priority = request.POST.get('priority', project.priority)
        project.budget = request.POST.get('budget', project.budget)
        project.progress = request.POST.get('progress', project.progress)
        
        user_id = request.POST.get('user_id')
        client_id = request.POST.get('client_id')
        if user_id:
            project.user = get_object_or_404(User, id=user_id)
        if client_id:
            project.client = get_object_or_404(Client, id=client_id)

        project.save()
        log_admin_activity(request.user, 'update', 'project', project.id, f"Updated project '{project.name}'", request)
        messages.success(request, f"Project '{project.name}' updated successfully.")
    return redirect('core:admin_projects_list')


@admin_required
def admin_project_archive(request, project_id):
    """Admin archive/restore project."""
    project = get_object_or_404(Project, id=project_id)
    project.is_archived = not project.is_archived
    project.save()
    status_str = "archived" if project.is_archived else "restored from archive"
    messages.success(request, f"Project '{project.name}' {status_str}.")
    return redirect('core:admin_projects_list')


@admin_required
def admin_project_duplicate(request, project_id):
    """Duplicate project."""
    project = get_object_or_404(Project, id=project_id)
    new_project = Project.objects.create(
        user=project.user,
        client=project.client,
        name=f"Copy of {project.name}",
        description=project.description,
        status='pending',
        priority=project.priority,
        budget=project.budget,
        progress=0
    )
    log_admin_activity(request.user, 'create', 'project', new_project.id, f"Duplicated project from '{project.name}'", request)
    messages.success(request, f"Duplicated project '{project.name}'.")
    return redirect('core:admin_projects_list')


@admin_required
def admin_project_delete(request, project_id):
    """Delete project."""
    project = get_object_or_404(Project, id=project_id)
    name = project.name
    project.delete()
    log_admin_activity(request.user, 'delete', 'project', project_id, f"Deleted project '{name}'", request)
    messages.success(request, f"Project '{name}' deleted.")
    return redirect('core:admin_projects_list')


@admin_required
def admin_project_bulk_action(request):
    """Bulk update or delete projects."""
    if request.method == 'POST':
        action = request.POST.get('bulk_action')
        project_ids = request.POST.getlist('project_ids')
        if not project_ids:
            messages.error(request, "No projects selected.")
            return redirect('core:admin_projects_list')

        queryset = Project.objects.filter(id__in=project_ids)
        count = queryset.count()

        if action == 'delete':
            queryset.delete()
            messages.success(request, f"Bulk deleted {count} projects.")
        elif action == 'archive':
            queryset.update(is_archived=True)
            messages.success(request, f"Bulk archived {count} projects.")
        elif action == 'complete':
            queryset.update(status='completed', progress=100)
            messages.success(request, f"Marked {count} projects as Completed.")

    return redirect('core:admin_projects_list')


# ============================================================
# FEATURE 6 — CLIENT MANAGEMENT (ADMIN SCOPE)
# ============================================================

@admin_required
def admin_clients_list(request):
    """Global Client Management."""
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')

    clients_qs = Client.objects.select_related('user').annotate(
        project_count=Count('projects')
    ).order_by('-created_at')

    if search_query:
        clients_qs = clients_qs.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(company__icontains=search_query)
        )

    if status_filter:
        clients_qs = clients_qs.filter(status=status_filter)

    paginator = Paginator(clients_qs, 15)
    page = request.GET.get('page')
    clients = paginator.get_page(page)

    all_users = User.objects.filter(is_active=True)

    context = {
        'clients': clients,
        'all_users': all_users,
        'search_query': search_query,
        'status_filter': status_filter,
        'page_title': 'Global Client Directory',
    }
    return render(request, 'admin_dashboard/clients/client_list.html', context)


@admin_required
def admin_client_create(request):
    """Admin create client."""
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        company = request.POST.get('company', '').strip()
        status = request.POST.get('status', 'active')

        user = get_object_or_404(User, id=user_id)
        client = Client.objects.create(
            user=user, name=name, email=email, phone=phone, company=company, status=status
        )
        log_admin_activity(request.user, 'create', 'client', client.id, f"Created client '{name}'", request)
        messages.success(request, f"Client '{name}' created successfully.")
    return redirect('core:admin_clients_list')


@admin_required
def admin_client_merge(request):
    """Merge duplicate clients into a primary client record."""
    if request.method == 'POST':
        primary_id = request.POST.get('primary_client_id')
        duplicate_id = request.POST.get('duplicate_client_id')

        if primary_id == duplicate_id:
            messages.error(request, "Primary and duplicate client cannot be the same.")
            return redirect('core:admin_clients_list')

        primary = get_object_or_404(Client, id=primary_id)
        duplicate = get_object_or_404(Client, id=duplicate_id)

        # Reassign projects & invoices to primary client
        Project.objects.filter(client=duplicate).update(client=primary)
        Invoice.objects.filter(client=duplicate).update(client=primary)
        Income.objects.filter(client=duplicate).update(client=primary)

        dup_name = duplicate.name
        duplicate.delete()

        log_admin_activity(request.user, 'update', 'client', primary.id, f"Merged duplicate client '{dup_name}' into '{primary.name}'", request)
        messages.success(request, f"Merged client '{dup_name}' into '{primary.name}'.")
    return redirect('core:admin_clients_list')


@admin_required
def admin_client_delete(request, client_id):
    """Delete client."""
    client = get_object_or_404(Client, id=client_id)
    name = client.name
    client.delete()
    log_admin_activity(request.user, 'delete', 'client', client_id, f"Deleted client '{name}'", request)
    messages.success(request, f"Client '{name}' deleted.")
    return redirect('core:admin_clients_list')


# ============================================================
# FEATURE 7 — FINANCIAL MANAGEMENT
# ============================================================

@admin_required
def admin_finances(request):
    """Global Financial Management & Refund Approvals."""
    payments = Payment.objects.select_related('user', 'project', 'project__client').order_by('-created_at')[:20]
    refund_requests = RefundRequest.objects.select_related('payment', 'user').order_by('-created_at')
    incomes = Income.objects.select_related('user', 'client', 'project').order_by('-date')[:15]
    expenses = Expense.objects.select_related('user', 'project').order_by('-date')[:15]

    total_revenue = (Payment.objects.filter(status='paid').aggregate(t=Sum('amount'))['t'] or 0) + (Income.objects.aggregate(t=Sum('amount'))['t'] or 0)
    total_expense = Expense.objects.aggregate(t=Sum('amount'))['t'] or 0
    pending_payments = Payment.objects.filter(status='pending').aggregate(t=Sum('amount'))['t'] or 0

    context = {
        'payments': payments,
        'refund_requests': refund_requests,
        'incomes': incomes,
        'expenses': expenses,
        'total_revenue': float(total_revenue),
        'total_expense': float(total_expense),
        'net_profit': float(total_revenue - total_expense),
        'pending_payments': float(pending_payments),
        'page_title': 'Financial Management',
    }
    return render(request, 'admin_dashboard/finances/finance_dashboard.html', context)


@admin_required
def admin_refund_action(request, refund_id):
    """Approve or reject refund requests."""
    refund = get_object_or_404(RefundRequest, id=refund_id)
    if request.method == 'POST':
        action = request.POST.get('action')  # 'approve' or 'reject'
        admin_notes = request.POST.get('admin_notes', '')

        if action == 'approve':
            refund.status = 'approved'
            refund.admin_notes = admin_notes
            refund.save()

            # Update associated payment status
            refund.payment.status = 'cancelled'
            refund.payment.save()
            log_admin_activity(request.user, 'update', 'payment', refund.payment.id, f"Approved refund of ${refund.amount}", request)
            messages.success(request, f"Refund request #${refund.id} APPROVED.")
        elif action == 'reject':
            refund.status = 'rejected'
            refund.admin_notes = admin_notes
            refund.save()
            messages.success(request, f"Refund request #${refund.id} REJECTED.")

    return redirect('core:admin_finances')


# ============================================================
# FEATURE 8 — NOTIFICATION CENTER & ANNOUNCEMENTS
# ============================================================

@admin_required
def admin_notifications(request):
    """Notification System & Announcement Builder."""
    announcements = Announcement.objects.select_related('created_by').order_by('-created_at')
    system_notifications = Notification.objects.select_related('user').order_by('-created_at')[:30]
    all_users = User.objects.filter(is_active=True)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        message = request.POST.get('message', '').strip()
        is_pinned = request.POST.get('is_pinned') == 'on'
        target_type = request.POST.get('target_type', 'all')
        selected_user_ids = request.POST.getlist('selected_users')

        announcement = Announcement.objects.create(
            title=title,
            message=message,
            is_pinned=is_pinned,
            target_type=target_type,
            created_by=request.user
        )

        if target_type == 'selected' and selected_user_ids:
            target_users = User.objects.filter(id__in=selected_user_ids)
            announcement.target_users.set(target_users)
            for u in target_users:
                Notification.objects.create(
                    user=u, title=f"Announcement: {title}", message=message, notification_type='announcement'
                )
        else:
            # Broadcast to all users
            for u in User.objects.filter(is_active=True):
                Notification.objects.create(
                    user=u, title=f"Announcement: {title}", message=message, notification_type='announcement'
                )

        log_admin_activity(request.user, 'create', 'user', announcement.id, f"Published announcement '{title}'", request)
        messages.success(request, "Announcement published successfully!")
        return redirect('core:admin_notifications')

    context = {
        'announcements': announcements,
        'system_notifications': system_notifications,
        'all_users': all_users,
        'page_title': 'Notification Center & Broadcasts',
    }
    return render(request, 'admin_dashboard/notifications/announcements.html', context)


@admin_required
def admin_announcement_delete(request, announcement_id):
    """Delete an announcement."""
    announcement = get_object_or_404(Announcement, id=announcement_id)
    announcement.delete()
    messages.success(request, "Announcement deleted.")
    return redirect('core:admin_notifications')


# ============================================================
# FEATURE 9 — ACTIVITY LOGS AUDIT HUB
# ============================================================

@admin_required
def admin_activity_logs(request):
    """System-wide audit trail."""
    search_q = request.GET.get('q', '').strip()
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')

    logs_qs = ActivityLog.objects.select_related('user').order_by('-timestamp')

    if search_q:
        logs_qs = logs_qs.filter(Q(description__icontains=search_q) | Q(ip_address__icontains=search_q))
    if action_filter:
        logs_qs = logs_qs.filter(action=action_filter)
    if user_filter:
        logs_qs = logs_qs.filter(user__id=user_filter)

    paginator = Paginator(logs_qs, 25)
    page = request.GET.get('page')
    logs = paginator.get_page(page)

    all_users = User.objects.filter(is_active=True)

    context = {
        'logs': logs,
        'all_users': all_users,
        'search_q': search_q,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'page_title': 'Activity Audit Logs',
    }
    return render(request, 'admin_dashboard/activity_logs/activity_list.html', context)


# ============================================================
# FEATURE 10 — SECURITY CENTER
# ============================================================

@admin_required
def admin_security(request):
    """Security Center: IP Blocking, Active Sessions & Failed Login Monitoring."""
    failed_logins = LoginHistory.objects.filter(status__in=['failed', 'blocked']).order_by('-timestamp')[:20]
    blocked_ips = BlockedIP.objects.select_related('blocked_by').order_by('-created_at')

    # Parse Active Sessions
    active_sessions = []
    for s in Session.objects.filter(expire_date__gte=timezone.now()):
        data = s.get_decoded()
        user_id = data.get('_auth_user_id')
        if user_id:
            try:
                u = User.objects.get(id=user_id)
                profile = getattr(u, 'profile', None)
                active_sessions.append({
                    'session_key': s.session_key,
                    'user': u,
                    'last_ip': profile.last_login_ip if profile else 'N/A',
                    'last_login': profile.last_login_at if profile else u.last_login,
                    'expire_date': s.expire_date
                })
            except User.DoesNotExist:
                pass

    if request.method == 'POST':
        action = request.POST.get('security_action')
        if action == 'block_ip':
            ip = request.POST.get('ip_address', '').strip()
            reason = request.POST.get('reason', '').strip()
            if ip:
                BlockedIP.objects.get_or_create(ip_address=ip, defaults={'reason': reason, 'blocked_by': request.user})
                log_admin_activity(request.user, 'update', 'user', request.user.id, f"Blocked IP address {ip}", request)
                messages.success(request, f"IP Address '{ip}' blocked.")
        elif action == 'terminate_session':
            key = request.POST.get('session_key')
            Session.objects.filter(session_key=key).delete()
            messages.success(request, "Session terminated.")

        return redirect('core:admin_security')

    context = {
        'failed_logins': failed_logins,
        'blocked_ips': blocked_ips,
        'active_sessions': active_sessions,
        'page_title': 'Security Center',
    }
    return render(request, 'admin_dashboard/security/security_center.html', context)


@admin_required
def admin_unblock_ip(request, ip_id):
    """Unblock an IP address."""
    blocked_ip = get_object_or_404(BlockedIP, id=ip_id)
    ip_str = blocked_ip.ip_address
    blocked_ip.delete()
    messages.success(request, f"Unblocked IP address {ip_str}.")
    return redirect('core:admin_security')


# ============================================================
# FEATURE 11 — SYSTEM SETTINGS
# ============================================================

@admin_required
def admin_settings(request):
    """Configure system-wide settings."""
    if request.method == 'POST':
        setting_keys = [
            'site_name', 'contact_email', 'currency_symbol', 'timezone',
            'theme_mode', 'maintenance_mode', 'session_timeout', 'password_min_length',
            'smtp_host', 'smtp_port', 'smtp_user', 'smtp_password', 'smtp_use_tls'
        ]
        for key in setting_keys:
            if key in request.POST:
                SystemSetting.set_setting(key, request.POST.get(key, ''))

        log_admin_activity(request.user, 'update', 'user', request.user.id, "Updated system settings", request)
        messages.success(request, "System settings updated successfully.")
        return redirect('core:admin_settings')

    settings_dict = {
        'site_name': SystemSetting.get_setting('site_name', 'FreelanceTrack'),
        'contact_email': SystemSetting.get_setting('contact_email', 'admin@freelancetrack.com'),
        'currency_symbol': SystemSetting.get_setting('currency_symbol', '$'),
        'timezone': SystemSetting.get_setting('timezone', 'UTC'),
        'theme_mode': SystemSetting.get_setting('theme_mode', 'dark'),
        'maintenance_mode': SystemSetting.get_setting('maintenance_mode', 'false'),
        'session_timeout': SystemSetting.get_setting('session_timeout', '60'),
        'password_min_length': SystemSetting.get_setting('password_min_length', '8'),
        'smtp_host': SystemSetting.get_setting('smtp_host', 'smtp.gmail.com'),
        'smtp_port': SystemSetting.get_setting('smtp_port', '587'),
        'smtp_user': SystemSetting.get_setting('smtp_user', ''),
        'smtp_use_tls': SystemSetting.get_setting('smtp_use_tls', 'true'),
    }

    context = {
        'settings_dict': settings_dict,
        'page_title': 'System Settings',
    }
    return render(request, 'admin_dashboard/settings/system_settings.html', context)


# ============================================================
# FEATURE 12 — DATABASE MANAGEMENT
# ============================================================

@admin_required
def admin_database(request):
    """Database Backup, Restore & Storage Metrics."""
    db_file_size = "N/A"
    db_path = connection.settings_dict.get('NAME')
    if db_path and os.path.exists(db_path):
        db_file_size = f"{os.path.getsize(db_path) / (1024 * 1024):.2f} MB"

    media_dir = getattr(connection.settings_dict, 'MEDIA_ROOT', os.path.join(os.getcwd(), 'media'))
    media_size = 0
    if os.path.exists(media_dir):
        for root, dirs, files in os.walk(media_dir):
            for f in files:
                media_size += os.path.getsize(os.path.join(root, f))
    media_file_size = f"{media_size / (1024 * 1024):.2f} MB"

    table_stats = [
        {'name': 'Users', 'count': User.objects.count()},
        {'name': 'Projects', 'count': Project.objects.count()},
        {'name': 'Clients', 'count': Client.objects.count()},
        {'name': 'Payments', 'count': Payment.objects.count()},
        {'name': 'Tasks', 'count': Task.objects.count()},
        {'name': 'Activity Logs', 'count': ActivityLog.objects.count()},
        {'name': 'Incomes', 'count': Income.objects.count()},
        {'name': 'Expenses', 'count': Expense.objects.count()},
        {'name': 'Invoices', 'count': Invoice.objects.count()},
        {'name': 'Files', 'count': ProjectFile.objects.count()},
    ]

    context = {
        'db_file_size': db_file_size,
        'media_file_size': media_file_size,
        'table_stats': table_stats,
        'page_title': 'Database Management & Backups',
    }
    return render(request, 'admin_dashboard/database/db_management.html', context)


@admin_required
def admin_database_backup(request):
    """Generate and download database backup JSON fixture."""
    buf = io.StringIO()
    call_command('dumpdata', indent=2, stdout=buf)
    json_data = buf.getvalue()

    filename = f"db_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    log_admin_activity(request.user, 'create', 'user', request.user.id, f"Downloaded database backup '{filename}'", request)
    return response


@admin_required
def admin_database_restore(request):
    """Restore database from uploaded fixture file."""
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        temp_path = os.path.join(os.getcwd(), 'temp_restore.json')
        with open(temp_path, 'wb+') as destination:
            for chunk in backup_file.chunks():
                destination.write(chunk)

        try:
            call_command('loaddata', temp_path)
            messages.success(request, "Database restored successfully from backup!")
            log_admin_activity(request.user, 'update', 'user', request.user.id, "Restored database from backup file", request)
        except Exception as e:
            messages.error(request, f"Failed to restore database: {str(e)}")
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    return redirect('core:admin_database')


# ============================================================
# FEATURE 13 — REPORTS & ANALYTICS HUB
# ============================================================

@admin_required
def admin_reports(request):
    """Reports & Analytics Hub."""
    context = {'page_title': 'Reports & Analytics Hub'}
    return render(request, 'admin_dashboard/reports/reports_hub.html', context)


@admin_required
def admin_report_export(request, report_type, export_format):
    """
    Consolidated Report Exporter for Users, Projects, Clients, Revenue, Expenses, Payments, Profit.
    Supports Excel (.xlsx) and PDF formats.
    """
    filename = f"{report_type}_report_{timezone.now().strftime('%Y%m%d')}"

    # Build Header & Rows Data based on report type
    headers = []
    data_rows = []

    if report_type == 'users':
        headers = ['Username', 'Email', 'Role', 'Status', 'Verified', 'Date Joined']
        for u in User.objects.select_related('profile').all():
            p = getattr(u, 'profile', None)
            data_rows.append([
                u.username, u.email,
                p.role if p else 'user',
                'Active' if u.is_active and not (p and p.is_suspended) else 'Suspended/Inactive',
                'Yes' if p and p.is_verified else 'No',
                u.date_joined.strftime('%Y-%m-%d')
            ])

    elif report_type == 'projects':
        headers = ['Project Name', 'Client', 'Freelancer', 'Status', 'Priority', 'Budget ($)', 'Created Date']
        for pr in Project.objects.select_related('client', 'user').all():
            data_rows.append([
                pr.name, pr.client.name, pr.user.username, pr.status, pr.priority, float(pr.budget or 0), pr.created_at.strftime('%Y-%m-%d')
            ])

    elif report_type == 'clients':
        headers = ['Client Name', 'Company', 'Email', 'Phone', 'Owner User', 'Projects Count']
        for c in Client.objects.select_related('user').annotate(pcount=Count('projects')).all():
            data_rows.append([
                c.name, c.company or '-', c.email, c.phone or '-', c.user.username, c.pcount
            ])

    elif report_type in ['revenue', 'payments']:
        headers = ['Payment ID', 'User', 'Project', 'Amount ($)', 'Status', 'Method', 'Paid Date']
        for pay in Payment.objects.select_related('user', 'project').all():
            data_rows.append([
                str(pay.id)[:8], pay.user.username, pay.project.name, float(pay.amount), pay.status, pay.payment_method, str(pay.paid_date or '-')
            ])

    elif report_type == 'expenses':
        headers = ['Expense Title', 'User', 'Category', 'Amount ($)', 'Date']
        for ex in Expense.objects.select_related('user').all():
            data_rows.append([
                ex.title, ex.user.username, ex.category, float(ex.amount), str(ex.date)
            ])

    else: # profit / summary
        headers = ['Metric', 'Value ($)']
        tot_rev = float((Payment.objects.filter(status='paid').aggregate(t=Sum('amount'))['t'] or 0) + (Income.objects.aggregate(t=Sum('amount'))['t'] or 0))
        tot_exp = float(Expense.objects.aggregate(t=Sum('amount'))['t'] or 0)
        data_rows = [
            ['Total Revenue', tot_rev],
            ['Total Expense', tot_exp],
            ['Net Profit', tot_rev - tot_exp]
        ]

    # EXPORT EXCEL
    if export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.title()

        # Styles
        header_fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in data_rows:
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # EXPORT PDF
    elif export_format == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#6F42C1'),
            spaceAfter=12
        )

        elements.append(Paragraph(f"FreelanceTrack — {report_type.title()} Report", title_style))
        elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 15))

        table_data = [headers] + [[str(val) for val in row] for row in data_rows]
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6F42C1')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        elements.append(t)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    return redirect('core:admin_reports')
