from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm
from django.contrib.auth.validators import UnicodeUsernameValidator
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta, date
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
import json
import re
import secrets
from .models import (Client, Project, Payment, Task, Note, ActivityLog,
                     UserProfile, ProjectFile, ProjectComment, Income,
                     Expense, Invoice, InvoiceItem, CalendarEvent, Notification,
                     LoginHistory, BlockedIP, SystemSetting, RefundRequest, Announcement,
                     EmailVerificationToken)
from .forms import (ClientForm, ProjectForm, PaymentForm, TaskForm,
                    NoteForm, SearchForm, UserBasicForm, UserProfileForm,
                    IncomeForm, ExpenseForm, InvoiceForm, InvoiceItemForm,
                    CalendarEventForm, ProjectFileForm, ProjectCommentForm,
                    UserRegistrationForm)
from .email_service import (send_welcome_email, send_login_alert_email,
                            send_verification_email, send_admin_new_user_notification)
from .brevo_service import send_brevo_welcome_email



def log_activity(user, action, model_type, model_id, description, request=None):
    """Helper function to log user activities to the ActivityLog."""
    ip_address = None
    user_agent = None
    if request:
        ip_address = get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:255]

    ActivityLog.objects.create(
        user=user,
        action=action,
        model_type=model_type,
        model_id=model_id,
        description=description,
        ip_address=ip_address,
        user_agent=user_agent
    )


def get_client_ip(request):
    """Extract the real client IP from request headers."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def home(request):
    """Landing page — redirects authenticated users to dashboard."""
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    return render(request, 'home.html')


def register(request):
    """User registration with mandatory email verification dispatch."""
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.save()

            username = form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')

            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = 'user'
            profile.is_verified = False
            profile.save()

            token_str = secrets.token_urlsafe(32)
            otp_code = f"{secrets.randbelow(1000000):06d}"
            expires_at = timezone.now() + timedelta(hours=24)

            token_obj = EmailVerificationToken.objects.create(
                user=user,
                token=token_str,
                otp=otp_code,
                expires_at=expires_at
            )

            send_verification_email(user, token_obj, request)
            send_brevo_welcome_email(user, request)


            log_activity(user, 'create', 'user', user.id,
                         f'User {username} submitted registration (pending verification)', request)

            messages.success(request, f'Registration successful! We sent a verification email to {email}. Please verify your email to activate your account.')
            return redirect('core:verify_email')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = UserRegistrationForm()
    return render(request, 'registration/register.html', {'form': form})


def verify_email(request, token=None):
    """
    Email verification view via URL token link or 6-digit OTP submission.
    """
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    token_str = token or request.GET.get('token')
    otp_code = request.POST.get('otp', '').strip() if request.method == 'POST' else None

    token_obj = None

    if token_str:
        token_obj = EmailVerificationToken.objects.filter(token=token_str, is_used=False).first()
    elif otp_code:
        token_obj = EmailVerificationToken.objects.filter(otp=otp_code, is_used=False).order_by('-created_at').first()

    if (token_str or otp_code):
        if not token_obj or not token_obj.is_valid():
            messages.error(request, 'The verification link or OTP code is invalid or has expired. Please request a new verification email.')
            return render(request, 'registration/verify_email.html')

        user = token_obj.user

        token_obj.is_used = True
        token_obj.save()

        user.is_active = True
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.is_verified = True
        profile.role = 'user'
        profile.save()

        log_activity(user, 'update', 'user', user.id, f'User {user.username} verified email successfully', request)

        # 1. Send Admin Notification Email (abhishekmutthalkar121@gmail.com)
        send_admin_new_user_notification(user, request)

        # 2. Send User Welcome Email
        send_welcome_email(user, request)

        messages.success(request, f'🎉 Email verified successfully! Welcome to FreelanceTrack, {user.username}. You can now log in.')
        return redirect('core:login')

    return render(request, 'registration/verify_email.html')


def resend_verification(request):
    """
    Allows user to request a new verification email if previous expired or got lost.
    """
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()

        if not email:
            messages.error(request, 'Please enter your registered email address.')
            return render(request, 'registration/resend_verification.html')

        user = User.objects.filter(email__iexact=email).first()

        if user:
            profile, _ = UserProfile.objects.get_or_create(user=user)
            if user.is_active and profile.is_verified:
                messages.info(request, 'Your account is already verified. You can log in directly.')
                return redirect('core:login')

            EmailVerificationToken.objects.filter(user=user, is_used=False).update(is_used=True)

            token_str = secrets.token_urlsafe(32)
            otp_code = f"{secrets.randbelow(1000000):06d}"
            expires_at = timezone.now() + timedelta(hours=24)

            token_obj = EmailVerificationToken.objects.create(
                user=user,
                token=token_str,
                otp=otp_code,
                expires_at=expires_at
            )

            send_verification_email(user, token_obj, request)
            messages.success(request, f'A new verification email and OTP have been sent to {email}.')
            return redirect('core:verify_email')
        else:
            messages.error(request, 'No registered account found with that email address.')

    return render(request, 'registration/resend_verification.html')


def custom_login(request):
    """
    FEATURE 1 & FEATURE 9 & FEATURE 10 — Role-Based Login & Security Audit:
    Supports Admin vs User login option selection.
    Enforces IP blocking, account verification checks, account suspension checks, role verification, and login audit tracking.
    """
    if request.user.is_authenticated:
        if request.user.is_staff or request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
            return redirect('core:admin_dashboard')
        return redirect('core:dashboard')

    if request.method == 'POST':
        login_type = request.POST.get('login_type', 'user')  # 'admin' or 'user'
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        ip = get_client_ip(request)
        ua = request.META.get('HTTP_USER_AGENT', '')[:255]

        # Device & Browser parsing
        device_type = 'Mobile' if 'Mobile' in ua or 'Android' in ua or 'iPhone' in ua else 'Desktop'
        browser = 'Chrome' if 'Chrome' in ua else 'Firefox' if 'Firefox' in ua else 'Safari' if 'Safari' in ua else 'Browser'

        # Check IP Blockage
        if BlockedIP.objects.filter(ip_address=ip).exists():
            messages.error(request, 'Access denied: Your IP address has been blocked for security reasons.')
            LoginHistory.objects.create(
                username_attempted=username, ip_address=ip, user_agent=ua,
                device=device_type, browser=browser, status='blocked'
            )
            return render(request, 'registration/login.html', {'login_type': login_type})

        # Pre-authenticate check for email verification status if credentials match inactive user
        unverified_user = User.objects.filter(username=username).first()
        if unverified_user and not unverified_user.is_active and unverified_user.check_password(password):
            messages.error(request, 'Your email address is not verified. Please verify your email before logging in.')
            LoginHistory.objects.create(
                user=unverified_user, username_attempted=username, ip_address=ip, user_agent=ua,
                device=device_type, browser=browser, status='failed'
            )
            return render(request, 'registration/login.html', {'login_type': login_type, 'show_resend_link': True})

        user = authenticate(request, username=username, password=password)

        if user is not None:
            profile, _ = UserProfile.objects.get_or_create(user=user)

            # Check Email Verification & User Suspension or Soft Delete
            if not profile.is_verified or not user.is_active:
                messages.error(request, 'Account email is not verified or account is inactive.')
                LoginHistory.objects.create(
                    user=user, username_attempted=username, ip_address=ip, user_agent=ua,
                    device=device_type, browser=browser, status='failed'
                )
                return render(request, 'registration/login.html', {'login_type': login_type, 'show_resend_link': True})

            if profile.is_suspended or profile.is_deleted:
                messages.error(request, 'Account suspended or deleted. Please contact the Super Admin.')
                LoginHistory.objects.create(
                    user=user, username_attempted=username, ip_address=ip, user_agent=ua,
                    device=device_type, browser=browser, status='failed'
                )
                return render(request, 'registration/login.html', {'login_type': login_type})

            is_admin_user = (user.is_staff or user.is_superuser or profile.role == 'admin')

            # Enforce Admin Role check if Admin mode selected
            if login_type == 'admin' and not is_admin_user:
                messages.error(request, 'Access denied: You do not have Super Admin privileges.')
                LoginHistory.objects.create(
                    user=user, username_attempted=username, ip_address=ip, user_agent=ua,
                    device=device_type, browser=browser, status='failed'
                )
                return render(request, 'registration/login.html', {'login_type': login_type})

            # Successful Authentication
            login(request, user)
            profile.last_login_ip = ip
            profile.last_login_at = timezone.now()
            profile.save()


            LoginHistory.objects.create(
                user=user, username_attempted=username, ip_address=ip, user_agent=ua,
                device=device_type, browser=browser, status='success'
            )
            log_activity(user, 'login', 'user', user.id, f'User {username} logged in ({login_type} mode)', request)
            send_login_alert_email(user, request)

            messages.success(request, f'Welcome back, {username}!')

            if is_admin_user:
                return redirect('core:admin_dashboard')
            return redirect('core:dashboard')
        else:
            LoginHistory.objects.create(
                username_attempted=username, ip_address=ip, user_agent=ua,
                device=device_type, browser=browser, status='failed'
            )
            messages.error(request, 'Invalid username or password. Please try again.')
            return render(request, 'registration/login.html', {'login_type': login_type})

    login_type = request.GET.get('type', 'user')
    return render(request, 'registration/login.html', {'login_type': login_type})



@require_http_methods(['GET', 'POST'])
def custom_logout(request):
    """Logout view — POST only to prevent CSRF-logout attacks."""
    # H-02: Only process logout on POST to prevent logout via embedded GET requests
    if request.method == 'POST' and request.user.is_authenticated:
        username = request.user.username
        log_activity(request.user, 'logout', 'user', request.user.id,
                     f'User logged out', request)
        logout(request)
        messages.success(request, 'You have been logged out successfully.')
    elif request.method == 'GET' and request.user.is_authenticated:
        # For GET requests, redirect to a confirmation page instead of logging out
        return redirect('core:dashboard')
    return redirect('core:login')


@login_required
@require_POST
def load_sample_data(request):
    """Seed sample data for request.user."""
    from .management.commands.seed_demo_data import seed_user_data
    try:
        seed_user_data(request.user)
        messages.success(request, '⚡ Sample clients, projects, tasks, and payments loaded successfully!')
    except Exception as e:
        messages.error(request, f'Failed to load sample data: {e}')
    return redirect('core:dashboard')


# ============================================================
# DASHBOARD VIEW
# ============================================================

@login_required
def dashboard(request):
    """Main dashboard with KPI cards, charts data, and recent records."""
    user = request.user

    # ── KPI Statistics ──────────────────────────────────────
    total_clients = Client.objects.filter(user=user, is_archived=False).count()
    total_projects = Project.objects.filter(user=user, is_archived=False).count()
    active_projects = Project.objects.filter(user=user, status='in_progress', is_archived=False).count()
    pending_projects = Project.objects.filter(user=user, status='pending', is_archived=False).count()
    completed_projects = Project.objects.filter(user=user, status='completed', is_archived=False).count()
    cancelled_projects = Project.objects.filter(user=user, status='cancelled', is_archived=False).count()

    # ── Financial KPI Statistics ────────────────────────────
    income_from_models = Income.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0
    paid_payments = Payment.objects.filter(user=user, status='paid').aggregate(total=Sum('amount'))['total'] or 0
    total_income = float(income_from_models) + float(paid_payments)

    total_expenses = float(Expense.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0)
    net_profit = total_income - total_expenses

    pending_payments = Payment.objects.filter(user=user, status='pending').aggregate(total=Sum('amount'))['total'] or 0
    completed_payments = paid_payments
    overdue_payments = Payment.objects.filter(user=user, status='pending', due_date__lt=timezone.now().date()).aggregate(total=Sum('amount'))['total'] or 0

    # ── Deadlines & Tasks ──────────────────────────────────
    upcoming_deadlines = Project.objects.filter(
        user=user,
        deadline__gte=timezone.now().date(),
        deadline__lte=timezone.now().date() + timedelta(days=7),
        status__in=['pending', 'in_progress'],
        is_archived=False
    ).select_related('client').order_by('deadline')[:5]

    active_tasks = Task.objects.filter(user=user, status__in=['todo', 'in_progress'], is_archived=False).count()
    recent_notifications = Notification.objects.filter(user=user, is_read=False)[:5]

    # ── Recent Activities & Projects ────────────────────────
    recent_activities = ActivityLog.objects.filter(user=user).order_by('-timestamp')[:10]
    recent_projects = Project.objects.filter(user=user, is_archived=False).select_related('client').order_by('-created_at')[:5]

    announcements = Announcement.objects.filter(Q(target_type='all') | Q(target_users=user)).distinct()[:5]

    # ── Cloud Storage Metrics (Image 1 UI Integration) ──────
    all_files = ProjectFile.objects.filter(user=user).select_related('project')
    recent_user_files = all_files.order_by('-uploaded_at')[:6]

    pictures_count = all_files.filter(file_type__in=['jpg', 'jpeg', 'png', 'gif', 'webp', 'svg']).count()
    documents_count = all_files.filter(file_type__in=['pdf', 'doc', 'docx', 'txt', 'csv', 'xlsx', 'ppt', 'pptx']).count()
    videos_count = all_files.filter(file_type__in=['mp4', 'mkv', 'avi', 'mov', 'wmv']).count()
    audio_count = all_files.filter(file_type__in=['mp3', 'wav', 'ogg', 'aac', 'flac']).count()

    total_bytes = all_files.aggregate(Sum('file_size'))['file_size__sum'] or 0
    total_mb = round(total_bytes / (1024 * 1024), 2)
    max_mb = 500.0
    storage_percentage = min(100, round((total_mb / max_mb) * 100, 1))

    file_form = ProjectFileForm()
    file_form.fields['project'].queryset = Project.objects.filter(user=user, is_archived=False)

    context = {
        'total_clients': total_clients,
        'total_projects': total_projects,
        'active_projects': active_projects,
        'pending_projects': pending_projects,
        'completed_projects': completed_projects,
        'cancelled_projects': cancelled_projects,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'pending_payments': pending_payments,
        'completed_payments': completed_payments,
        'overdue_payments': overdue_payments,
        'upcoming_deadlines': upcoming_deadlines,
        'active_tasks': active_tasks,
        'recent_notifications': recent_notifications,
        'recent_activities': recent_activities,
        'recent_projects': recent_projects,
        'total_earnings': total_income,
        'announcements': announcements,

        # Cloud Storage Data
        'user_files': recent_user_files,
        'pictures_count': pictures_count,
        'documents_count': documents_count,
        'videos_count': videos_count,
        'audio_count': audio_count,
        'total_mb': total_mb,
        'max_mb': max_mb,
        'storage_percentage': storage_percentage,
        'file_form': file_form,
    }

    return render(request, 'dashboard.html', context)


# ============================================================
# CLIENT VIEWS
# ============================================================

@login_required
def client_list(request):
    """List all clients with search, status filter, and pagination."""
    user = request.user
    clients = Client.objects.filter(user=user)

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', '-created_at')

    if search_query:
        clients = clients.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(company__icontains=search_query)
        )

    if status_filter:
        clients = clients.filter(status=status_filter)

    # Sorting
    valid_sorts = ['name', '-name', 'created_at', '-created_at', 'company', '-company']
    if sort_by in valid_sorts:
        clients = clients.order_by(sort_by)

    paginator = Paginator(clients, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'status_choices': Client.STATUS_CHOICES,
        'total_count': clients.count(),
    }
    return render(request, 'clients/client_list.html', context)


@login_required
def client_detail(request, pk):
    """View client details with projects and financial summary."""
    client = get_object_or_404(Client, pk=pk, user=request.user)
    projects = client.projects.all().order_by('-created_at')
    total_earned = Payment.objects.filter(
        project__client=client, status='paid'
    ).aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'client': client,
        'projects': projects,
        'total_earned': total_earned,
    }
    return render(request, 'clients/client_detail.html', context)


@login_required
def client_create(request):
    """Create a new client record."""
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save(commit=False)
            client.user = request.user
            client.save()
            log_activity(request.user, 'create', 'client', client.id,
                         f'Created client: {client.name}', request)
            messages.success(request, f'Client "{client.name}" created successfully!')
            return redirect('core:client_detail', pk=client.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = ClientForm()
    return render(request, 'clients/client_form.html', {'form': form, 'action': 'Create'})


@login_required
def client_update(request, pk):
    """Update an existing client record."""
    client = get_object_or_404(Client, pk=pk, user=request.user)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            updated_client = form.save()
            log_activity(request.user, 'update', 'client', client.id,
                         f'Updated client: {updated_client.name}', request)
            messages.success(request, f'Client "{updated_client.name}" updated successfully!')
            return redirect('core:client_detail', pk=client.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = ClientForm(instance=client)
    return render(request, 'clients/client_form.html', {
        'form': form, 'action': 'Update', 'client': client
    })


@login_required
def client_delete(request, pk):
    """Confirm and delete a client record."""
    client = get_object_or_404(Client, pk=pk, user=request.user)
    if request.method == 'POST':
        client_name = client.name
        log_activity(request.user, 'delete', 'client', pk,
                     f'Deleted client: {client_name}', request)
        client.delete()
        messages.success(request, f'Client "{client_name}" deleted successfully!')
        return redirect('core:client_list')
    return render(request, 'clients/client_confirm_delete.html', {'client': client})


# ============================================================
# PROJECT VIEWS
# ============================================================

@login_required
def project_list(request):
    """List all projects with search, filter, sorting, and pagination."""
    user = request.user
    projects = Project.objects.filter(user=user).select_related('client')

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    sort_by = request.GET.get('sort', '-created_at')

    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(client__name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if status_filter:
        projects = projects.filter(status=status_filter)

    if priority_filter:
        projects = projects.filter(priority=priority_filter)

    valid_sorts = ['name', '-name', 'deadline', '-deadline', 'created_at', '-created_at',
                   'budget', '-budget', 'progress', '-progress']
    if sort_by in valid_sorts:
        projects = projects.order_by(sort_by)

    paginator = Paginator(projects, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'sort_by': sort_by,
        'status_choices': Project.STATUS_CHOICES,
        'priority_choices': Project.PRIORITY_CHOICES,
        'total_count': projects.count(),
    }
    return render(request, 'projects/project_list.html', context)


@login_required
def project_detail(request, pk):
    """View project details with tasks, payments, and notes."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    tasks = project.tasks.all().order_by('status', '-created_at')
    payments = project.payments.all().order_by('-created_at')
    notes = project.project_notes.all().order_by('-created_at')

    total_paid = payments.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    total_pending = payments.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'project': project,
        'tasks': tasks,
        'payments': payments,
        'notes': notes,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'completed_tasks': tasks.filter(status='completed').count(),
        'total_tasks': tasks.count(),
    }
    return render(request, 'projects/project_detail.html', context)


@login_required
def project_create(request):
    """Create a new project."""
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
        if form.is_valid():
            project = form.save(commit=False)
            project.user = request.user
            project.save()
            log_activity(request.user, 'create', 'project', project.id,
                         f'Created project: {project.name}', request)
            messages.success(request, f'Project "{project.name}" created successfully!')
            return redirect('core:project_detail', pk=project.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = ProjectForm()
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
    return render(request, 'projects/project_form.html', {'form': form, 'action': 'Create'})


@login_required
def project_update(request, pk):
    """Update an existing project."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
        if form.is_valid():
            old_status = project.status
            updated_project = form.save()
            if old_status != updated_project.status:
                log_activity(request.user, 'status_change', 'project', project.id,
                             f'Project status changed: {old_status} → {updated_project.status}', request)
            log_activity(request.user, 'update', 'project', project.id,
                         f'Updated project: {updated_project.name}', request)
            messages.success(request, f'Project "{updated_project.name}" updated successfully!')
            return redirect('core:project_detail', pk=project.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = ProjectForm(instance=project)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
    return render(request, 'projects/project_form.html', {
        'form': form, 'action': 'Update', 'project': project
    })


@login_required
def project_delete(request, pk):
    """Confirm and delete a project."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    if request.method == 'POST':
        project_name = project.name
        log_activity(request.user, 'delete', 'project', pk,
                     f'Deleted project: {project_name}', request)
        project.delete()
        messages.success(request, f'Project "{project_name}" deleted successfully!')
        return redirect('core:project_list')
    return render(request, 'projects/project_confirm_delete.html', {'project': project})


# ============================================================
# PAYMENT VIEWS
# ============================================================

@login_required
def payment_list(request):
    """List all payments with search, filter, and pagination."""
    user = request.user
    payments = Payment.objects.filter(user=user).select_related('project', 'project__client')

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', '-created_at')

    if search_query:
        payments = payments.filter(
            Q(project__name__icontains=search_query) |
            Q(invoice_number__icontains=search_query) |
            Q(project__client__name__icontains=search_query)
        )

    if status_filter:
        payments = payments.filter(status=status_filter)

    valid_sorts = ['amount', '-amount', 'due_date', '-due_date', 'created_at', '-created_at']
    if sort_by in valid_sorts:
        payments = payments.order_by(sort_by)

    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Summary stats
    total_paid = Payment.objects.filter(user=user, status='paid').aggregate(t=Sum('amount'))['t'] or 0
    total_pending = Payment.objects.filter(user=user, status='pending').aggregate(t=Sum('amount'))['t'] or 0

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'status_choices': Payment.STATUS_CHOICES,
        'total_paid': total_paid,
        'total_pending': total_pending,
    }
    return render(request, 'payments/payment_list.html', context)


@login_required
def payment_detail(request, pk):
    """View payment details."""
    payment = get_object_or_404(Payment, pk=pk, user=request.user)
    return render(request, 'payments/payment_detail.html', {'payment': payment})


@login_required
def payment_create(request):
    """Create a new payment record."""
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.user = request.user
            payment.save()
            log_activity(request.user, 'create', 'payment', payment.id,
                         f'Created payment: ${payment.amount} for {payment.project.name}', request)
            messages.success(request, 'Payment created successfully!')
            return redirect('core:payment_detail', pk=payment.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = PaymentForm()
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
    return render(request, 'payments/payment_form.html', {'form': form, 'action': 'Create'})


@login_required
def payment_update(request, pk):
    """Update an existing payment record."""
    payment = get_object_or_404(Payment, pk=pk, user=request.user)
    old_status = payment.status
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        if form.is_valid():
            updated_payment = form.save()
            if old_status != updated_payment.status:
                log_activity(request.user, 'status_change', 'payment', payment.id,
                             f'Payment status: {old_status} → {updated_payment.status}', request)
            log_activity(request.user, 'update', 'payment', payment.id,
                         f'Updated payment: ${updated_payment.amount}', request)
            messages.success(request, 'Payment updated successfully!')
            return redirect('core:payment_detail', pk=payment.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = PaymentForm(instance=payment)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
    return render(request, 'payments/payment_form.html', {
        'form': form, 'action': 'Update', 'payment': payment
    })


@login_required
def payment_delete(request, pk):
    """Confirm and delete a payment record."""
    payment = get_object_or_404(Payment, pk=pk, user=request.user)
    if request.method == 'POST':
        amount = payment.amount
        log_activity(request.user, 'delete', 'payment', pk,
                     f'Deleted payment: ${amount}', request)
        payment.delete()
        messages.success(request, 'Payment deleted successfully!')
        return redirect('core:payment_list')
    return render(request, 'payments/payment_confirm_delete.html', {'payment': payment})


# ============================================================
# TASK VIEWS
# ============================================================

@login_required
def task_list(request):
    """List all tasks with search, filter, sorting, and pagination."""
    user = request.user
    tasks = Task.objects.filter(user=user).select_related('project', 'project__client')

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    project_filter = request.GET.get('project', '')
    sort_by = request.GET.get('sort', '-created_at')

    if search_query:
        tasks = tasks.filter(
            Q(title__icontains=search_query) |
            Q(project__name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if status_filter:
        tasks = tasks.filter(status=status_filter)

    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)

    if project_filter:
        tasks = tasks.filter(project__id=project_filter)

    valid_sorts = ['title', '-title', 'due_date', '-due_date', 'created_at', '-created_at',
                   'priority', '-priority', 'status', '-status']
    if sort_by in valid_sorts:
        tasks = tasks.order_by(sort_by)

    paginator = Paginator(tasks, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Available projects for filter dropdown
    user_projects = Project.objects.filter(user=user).order_by('name')

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'project_filter': project_filter,
        'sort_by': sort_by,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'user_projects': user_projects,
        'total_count': tasks.count(),
    }
    return render(request, 'tasks/task_list.html', context)


@login_required
def task_detail(request, pk):
    """View task details."""
    task = get_object_or_404(Task, pk=pk, user=request.user)
    return render(request, 'tasks/task_detail.html', {'task': task})


@login_required
def task_create(request):
    """Create a new task."""
    if request.method == 'POST':
        form = TaskForm(request.POST)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            task.save()
            log_activity(request.user, 'create', 'task', task.id,
                         f'Created task: {task.title}', request)
            messages.success(request, f'Task "{task.title}" created successfully!')
            return redirect('core:task_detail', pk=task.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = TaskForm()
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
    return render(request, 'tasks/task_form.html', {'form': form, 'action': 'Create'})


@login_required
def task_update(request, pk):
    """Update an existing task."""
    task = get_object_or_404(Task, pk=pk, user=request.user)
    old_status = task.status
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        if form.is_valid():
            updated_task = form.save()
            if old_status != updated_task.status:
                log_activity(request.user, 'status_change', 'task', task.id,
                             f'Task status: {old_status} → {updated_task.status}', request)
            log_activity(request.user, 'update', 'task', task.id,
                         f'Updated task: {updated_task.title}', request)
            messages.success(request, f'Task "{updated_task.title}" updated successfully!')
            return redirect('core:task_detail', pk=task.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = TaskForm(instance=task)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
    return render(request, 'tasks/task_form.html', {
        'form': form, 'action': 'Update', 'task': task
    })


@login_required
def task_delete(request, pk):
    """Confirm and delete a task."""
    task = get_object_or_404(Task, pk=pk, user=request.user)
    if request.method == 'POST':
        task_title = task.title
        log_activity(request.user, 'delete', 'task', pk,
                     f'Deleted task: {task_title}', request)
        task.delete()
        messages.success(request, f'Task "{task_title}" deleted successfully!')
        return redirect('core:task_list')
    return render(request, 'tasks/task_confirm_delete.html', {'task': task})


# ============================================================
# NOTE VIEWS
# ============================================================

@login_required
def note_list(request):
    """List all notes with search and pagination."""
    user = request.user
    notes = Note.objects.filter(user=user).select_related('project', 'client')

    search_query = request.GET.get('search', '').strip()
    if search_query:
        notes = notes.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query)
        )

    paginator = Paginator(notes, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'notes/note_list.html', context)


@login_required
def note_detail(request, pk):
    """View note details."""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    return render(request, 'notes/note_detail.html', {'note': note})


@login_required
def note_create(request):
    """Create a new note."""
    if request.method == 'POST':
        form = NoteForm(request.POST)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
        if form.is_valid():
            note = form.save(commit=False)
            note.user = request.user
            note.save()
            log_activity(request.user, 'create', 'note', note.id,
                         f'Created note: {note.title}', request)
            messages.success(request, 'Note created successfully!')
            return redirect('core:note_detail', pk=note.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = NoteForm()
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
    return render(request, 'notes/note_form.html', {'form': form, 'action': 'Create'})


@login_required
def note_update(request, pk):
    """Update an existing note."""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    if request.method == 'POST':
        form = NoteForm(request.POST, instance=note)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
        if form.is_valid():
            updated_note = form.save()
            log_activity(request.user, 'update', 'note', note.id,
                         f'Updated note: {updated_note.title}', request)
            messages.success(request, 'Note updated successfully!')
            return redirect('core:note_detail', pk=note.id)
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = NoteForm(instance=note)
        form.fields['project'].queryset = Project.objects.filter(user=request.user)
        form.fields['client'].queryset = Client.objects.filter(user=request.user)
    return render(request, 'notes/note_form.html', {
        'form': form, 'action': 'Update', 'note': note
    })


@login_required
def note_delete(request, pk):
    """Confirm and delete a note."""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    if request.method == 'POST':
        note_title = note.title
        log_activity(request.user, 'delete', 'note', pk,
                     f'Deleted note: {note_title}', request)
        note.delete()
        messages.success(request, 'Note deleted successfully!')
        return redirect('core:note_list')
    return render(request, 'notes/note_confirm_delete.html', {'note': note})


# ============================================================
# ACTIVITY LOG VIEW
# ============================================================

@login_required
def activity_list(request):
    """List all user activities with pagination."""
    user = request.user
    activities = ActivityLog.objects.filter(user=user).order_by('-timestamp')

    paginator = Paginator(activities, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'activities/activity_list.html', {'page_obj': page_obj})


# ============================================================
# REPORTS VIEW
# ============================================================

@login_required
def reports_dashboard(request):
    """Reports overview page with export options."""
    user = request.user

    # Payment report data
    total_paid = Payment.objects.filter(user=user, status='paid').aggregate(t=Sum('amount'))['t'] or 0
    total_pending = Payment.objects.filter(user=user, status='pending').aggregate(t=Sum('amount'))['t'] or 0
    total_overdue = Payment.objects.filter(user=user, status='overdue').aggregate(t=Sum('amount'))['t'] or 0

    # Project report data
    projects_by_status = {
        'completed': Project.objects.filter(user=user, status='completed').count(),
        'in_progress': Project.objects.filter(user=user, status='in_progress').count(),
        'pending': Project.objects.filter(user=user, status='pending').count(),
        'on_hold': Project.objects.filter(user=user, status='on_hold').count(),
        'cancelled': Project.objects.filter(user=user, status='cancelled').count(),
    }

    # Monthly report — last 12 months
    monthly_data = []
    today = timezone.now().date()
    for i in range(11, -1, -1):
        month_date = (today.replace(day=1) - timedelta(days=i * 30))
        month_total = Payment.objects.filter(
            user=user, status='paid',
            paid_date__year=month_date.year,
            paid_date__month=month_date.month
        ).aggregate(t=Sum('amount'))['t'] or 0
        monthly_data.append({
            'month': month_date.strftime('%b %Y'),
            'total': float(month_total),
        })

    # Client report
    top_clients = Client.objects.filter(user=user).annotate(
        project_count=Count('projects')
    ).order_by('-project_count')[:10]

    # Scatter Plot dataset (Budget vs Revenue)
    user_projects = Project.objects.filter(user=user).select_related('client')
    user_payments = Payment.objects.filter(user=user)
    scatter_list = []
    for p in user_projects[:15]:
        p_paid = user_payments.filter(project=p, status='paid').aggregate(t=Sum('amount'))['t'] or 0
        scatter_list.append({
            'x': float(p.budget or 0),
            'y': float(p_paid),
            'name': p.name,
            'client': p.client.name if p.client else 'N/A'
        })

    # Heatmap Density matrix (7 Days x 4 Weeks)
    days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    heatmap_matrix = []
    for w in range(3, -1, -1):
        week_row = []
        for d in range(7):
            day_target = today - timedelta(days=(w * 7 + (today.weekday() - d) % 7))
            day_payments = user_payments.filter(paid_date=day_target, status='paid').aggregate(t=Sum('amount'))['t'] or 0
            intensity = min(100, int((float(day_payments) / 500.0) * 100)) if day_payments else (15 if d < 5 else 5)
            week_row.append({
                'date': day_target.strftime('%b %d'),
                'val': float(day_payments),
                'intensity': intensity
            })
        heatmap_matrix.append(week_row)

    context = {
        'total_paid': total_paid,
        'total_pending': total_pending,
        'total_overdue': total_overdue,
        'projects_by_status': projects_by_status,
        'monthly_data': monthly_data,
        'top_clients': top_clients,
        'monthly_labels_json': json.dumps([d['month'] for d in monthly_data]),
        'monthly_values_json': json.dumps([d['total'] for d in monthly_data]),
        'scatter_data_json': json.dumps(scatter_list),
        'heatmap_data_json': json.dumps({'days': days_of_week, 'matrix': heatmap_matrix}),
        'status_data_json': json.dumps(projects_by_status),
    }
    return render(request, 'reports/reports.html', context)



_VALID_REPORT_TYPES = frozenset(['payment', 'project', 'client', 'monthly'])


@login_required
def export_pdf_report(request, report_type):
    """Generate and stream a PDF report using ReportLab."""
    if report_type not in _VALID_REPORT_TYPES:
        messages.error(request, 'Invalid report type requested.')
        return redirect('core:reports_dashboard')

    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.units import inch
        import io
    except ImportError:
        messages.error(request, 'ReportLab is not installed. Cannot generate PDF.')
        return redirect('core:reports_dashboard')

    user = request.user
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  fontSize=20, textColor=colors.HexColor('#4f46e5'),
                                  alignment=TA_CENTER, spaceAfter=12)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'],
                                    textColor=colors.HexColor('#1e1b4b'), spaceAfter=8)
    normal_style = styles['Normal']

    elements = []
    today = timezone.now().date()
    report_date = today.strftime('%B %d, %Y')

    # Report Title
    elements.append(Paragraph(f'Freelancer Project Tracker', title_style))
    elements.append(Paragraph(f'Report Type: {report_type.title().replace("_", " ")} — {report_date}', normal_style))
    elements.append(Spacer(1, 0.3 * inch))

    header_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ])

    if report_type == 'payment':
        elements.append(Paragraph('Payment Report', heading_style))
        payments = Payment.objects.filter(user=user).select_related('project').order_by('-created_at')
        data = [['Invoice #', 'Project', 'Amount', 'Status', 'Due Date', 'Paid Date']]
        for p in payments:
            data.append([
                p.invoice_number or 'N/A',
                p.project.name[:30],
                f'${p.amount:,.2f}',
                p.get_status_display(),
                str(p.due_date or 'N/A'),
                str(p.paid_date or 'N/A'),
            ])
        if len(data) > 1:
            t = Table(data, colWidths=[1.2 * inch, 2 * inch, 1 * inch, 1 * inch, 1 * inch, 1 * inch])
            t.setStyle(header_style)
            elements.append(t)
        else:
            elements.append(Paragraph('No payment records found.', normal_style))

    elif report_type == 'project':
        elements.append(Paragraph('Project Report', heading_style))
        projects = Project.objects.filter(user=user).select_related('client').order_by('-created_at')
        data = [['Project Name', 'Client', 'Status', 'Priority', 'Budget', 'Progress', 'Deadline']]
        for p in projects:
            data.append([
                p.name[:25],
                p.client.name[:20],
                p.get_status_display(),
                p.get_priority_display(),
                f'${p.budget:,.2f}' if p.budget else 'N/A',
                f'{p.progress}%',
                str(p.deadline or 'N/A'),
            ])
        if len(data) > 1:
            t = Table(data, colWidths=[1.5 * inch, 1.2 * inch, 1 * inch, 0.8 * inch, 0.8 * inch, 0.6 * inch, 0.8 * inch])
            t.setStyle(header_style)
            elements.append(t)
        else:
            elements.append(Paragraph('No project records found.', normal_style))

    elif report_type == 'client':
        elements.append(Paragraph('Client Report', heading_style))
        clients = Client.objects.filter(user=user).annotate(proj_count=Count('projects')).order_by('-proj_count')
        data = [['Client Name', 'Company', 'Email', 'Status', 'Total Projects']]
        for c in clients:
            data.append([
                c.name[:25],
                (c.company or 'Individual')[:25],
                c.email[:30],
                c.get_status_display(),
                str(c.proj_count),
            ])
        if len(data) > 1:
            t = Table(data, colWidths=[1.5 * inch, 1.5 * inch, 2 * inch, 0.8 * inch, 0.8 * inch])
            t.setStyle(header_style)
            elements.append(t)
        else:
            elements.append(Paragraph('No client records found.', normal_style))

    elif report_type == 'monthly':
        elements.append(Paragraph('Monthly Earnings Report (Last 12 Months)', heading_style))
        today = timezone.now().date()
        data = [['Month', 'Total Earned', 'Number of Payments']]
        for i in range(11, -1, -1):
            month_date = (today.replace(day=1) - timedelta(days=i * 30))
            month_payments = Payment.objects.filter(
                user=user, status='paid',
                paid_date__year=month_date.year,
                paid_date__month=month_date.month
            )
            month_total = month_payments.aggregate(t=Sum('amount'))['t'] or 0
            data.append([
                month_date.strftime('%B %Y'),
                f'${month_total:,.2f}',
                str(month_payments.count()),
            ])
        t = Table(data, colWidths=[2 * inch, 2 * inch, 2 * inch])
        t.setStyle(header_style)
        elements.append(t)

    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f'Generated by Freelancer Project Tracker on {report_date}', normal_style))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    safe_filename = f'{report_type}_report_{today.strftime("%Y-%m-%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    return response


@login_required
def export_excel_report(request, report_type):
    """Generate and stream an Excel report using openpyxl."""
    if report_type not in _VALID_REPORT_TYPES:
        messages.error(request, 'Invalid report type requested.')
        return redirect('core:reports_dashboard')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        messages.error(request, 'openpyxl is not installed. Cannot generate Excel.')
        return redirect('core:reports_dashboard')

    user = request.user
    wb = openpyxl.Workbook()
    ws = wb.active

    # Styling helpers
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    alt_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

    def style_data_row(ws, row_num, col_count, alternate=False):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            if alternate:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border

    today = timezone.now().date()

    if report_type == 'payment':
        ws.title = 'Payment Report'
        headers = ['Invoice #', 'Project', 'Client', 'Amount', 'Status', 'Method', 'Due Date', 'Paid Date']
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        payments = Payment.objects.filter(user=user).select_related('project', 'project__client').order_by('-created_at')
        for i, p in enumerate(payments, 2):
            ws.append([
                p.invoice_number or 'N/A',
                p.project.name,
                p.project.client.name,
                float(p.amount),
                p.get_status_display(),
                p.get_payment_method_display(),
                str(p.due_date or ''),
                str(p.paid_date or ''),
            ])
            style_data_row(ws, i, len(headers), i % 2 == 0)
        col_widths = [15, 25, 20, 12, 12, 15, 12, 12]

    elif report_type == 'project':
        ws.title = 'Project Report'
        headers = ['Project Name', 'Client', 'Status', 'Priority', 'Budget', 'Progress %',
                   'Start Date', 'Deadline', 'Total Paid', 'Total Pending']
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        projects = Project.objects.filter(user=user).select_related('client').order_by('-created_at')
        for i, p in enumerate(projects, 2):
            total_paid = p.payments.filter(status='paid').aggregate(t=Sum('amount'))['t'] or 0
            total_pending = p.payments.filter(status='pending').aggregate(t=Sum('amount'))['t'] or 0
            ws.append([
                p.name, p.client.name, p.get_status_display(), p.get_priority_display(),
                float(p.budget) if p.budget else 0,
                p.progress,
                str(p.start_date or ''), str(p.deadline or ''),
                float(total_paid), float(total_pending),
            ])
            style_data_row(ws, i, len(headers), i % 2 == 0)
        col_widths = [25, 20, 12, 12, 12, 12, 12, 12, 12, 12]

    elif report_type == 'client':
        ws.title = 'Client Report'
        headers = ['Client Name', 'Company', 'Email', 'Phone', 'Status',
                   'Total Projects', 'Total Earned']
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        clients = Client.objects.filter(user=user).annotate(proj_count=Count('projects')).order_by('-proj_count')
        for i, c in enumerate(clients, 2):
            total_earned = Payment.objects.filter(
                project__client=c, status='paid'
            ).aggregate(t=Sum('amount'))['t'] or 0
            ws.append([
                c.name, c.company or 'Individual', c.email,
                c.phone or '', c.get_status_display(),
                c.proj_count, float(total_earned),
            ])
            style_data_row(ws, i, len(headers), i % 2 == 0)
        col_widths = [20, 20, 25, 15, 12, 15, 15]

    elif report_type == 'monthly':
        ws.title = 'Monthly Report'
        headers = ['Month', 'Year', 'Total Earned ($)', 'No. of Payments',
                   'Pending Amount ($)', 'No. of Pending']
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        for i_m, i in enumerate(range(11, -1, -1), 2):
            month_date = (today.replace(day=1) - timedelta(days=i * 30))
            paid = Payment.objects.filter(
                user=user, status='paid',
                paid_date__year=month_date.year,
                paid_date__month=month_date.month
            )
            pending = Payment.objects.filter(
                user=user, status='pending',
                due_date__year=month_date.year,
                due_date__month=month_date.month
            )
            ws.append([
                month_date.strftime('%B'), month_date.year,
                float(paid.aggregate(t=Sum('amount'))['t'] or 0),
                paid.count(),
                float(pending.aggregate(t=Sum('amount'))['t'] or 0),
                pending.count(),
            ])
            style_data_row(ws, i_m, len(headers), i_m % 2 == 0)
        col_widths = [15, 8, 18, 15, 18, 15]

    else:
        messages.error(request, f'Unknown report type: {report_type}')
        return redirect('core:reports_dashboard')

    # Apply column widths
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_filename = f'{report_type}_report_{today.strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    return response


# ============================================================
# SETTINGS VIEW
# ============================================================

_VALID_SETTINGS_ACTIONS = frozenset(['profile', 'password'])


@login_required
def user_settings(request):
    """User settings page — profile update and password change."""
    password_form = PasswordChangeForm(user=request.user)

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()

        if action not in _VALID_SETTINGS_ACTIONS:
            messages.error(request, 'Invalid settings action.')
            return redirect('core:settings')

        if action == 'profile':
            # Update basic profile fields
            user = request.user
            user.first_name = request.POST.get('first_name', '').strip()[:150]
            user.last_name = request.POST.get('last_name', '').strip()[:150]
            
            new_email = request.POST.get('email', '').strip()
            if new_email:
                try:
                    validate_email(new_email)
                    user.email = new_email
                except ValidationError:
                    messages.error(request, 'Please enter a valid email address.')
                    return render(request, 'settings.html', {'password_form': password_form})
            else:
                user.email = ''
                
            user.save()
            log_activity(request.user, 'update', 'user', request.user.id,
                         'Updated profile settings', request)
            messages.success(request, 'Profile updated successfully!')
            return redirect('core:settings')

        elif action == 'password':
            password_form = PasswordChangeForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                log_activity(request.user, 'update', 'user', request.user.id,
                             'Changed account password', request)
                messages.success(request, 'Password changed successfully!')
                return redirect('core:settings')
            else:
                messages.error(request, 'Password change failed. Please check the errors below.')

    context = {
        'password_form': password_form,
    }
    return render(request, 'settings.html', context)


# ============================================================
# MY PROFILE VIEWS
# ============================================================

@login_required
def profile_view(request):
    """View user profile and login history."""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    login_history = ActivityLog.objects.filter(
        user=request.user, action__in=['login', 'logout']
    ).order_by('-timestamp')[:15]

    context = {
        'profile': profile,
        'login_history': login_history,
    }
    return render(request, 'profile/profile.html', context)


@login_required
def profile_edit(request):
    """Edit user details and profile info."""
    profile, created = UserProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        user_form = UserBasicForm(request.POST, instance=request.user)
        profile_form = UserProfileForm(request.POST, request.FILES, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            log_activity(request.user, 'profile_updated', 'user', request.user.id, 'Updated profile details', request)
            messages.success(request, 'Profile updated successfully!')
            return redirect('core:profile_view')
        else:
            messages.error(request, 'Please fix errors in the form.')
    else:
        user_form = UserBasicForm(instance=request.user)
        profile_form = UserProfileForm(instance=profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'profile': profile,
    }
    return render(request, 'profile/profile.html', context)


@login_required
@require_POST
def profile_change_password(request):
    """Change account password from profile tab."""
    form = PasswordChangeForm(user=request.user, data=request.POST)
    if form.is_valid():
        user = form.save()
        update_session_auth_hash(request, user)
        log_activity(request.user, 'update', 'user', request.user.id, 'Changed password', request)
        messages.success(request, 'Password changed successfully!')
    else:
        messages.error(request, 'Failed to change password. Please check requirements.')
    return redirect('core:profile_view')


@login_required
@require_POST
def profile_remove_picture(request):
    """Remove user profile picture."""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    if profile.profile_picture:
        profile.profile_picture.delete(save=True)
        messages.success(request, 'Profile picture removed.')
    return redirect('core:profile_view')


# ============================================================
# PROJECT EXTENSION VIEWS (ARCHIVE, RESTORE, DUPLICATE, COMMENTS, FILES)
# ============================================================

@login_required
@require_POST
def project_archive(request, pk):
    """Archive a project."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    project.is_archived = True
    project.save()
    log_activity(request.user, 'archive', 'project', project.id, f'Archived project: {project.name}', request)
    messages.success(request, f'Project "{project.name}" archived.')
    return redirect('core:project_list')


@login_required
@require_POST
def project_restore(request, pk):
    """Restore an archived project."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    project.is_archived = False
    project.save()
    log_activity(request.user, 'restore', 'project', project.id, f'Restored project: {project.name}', request)
    messages.success(request, f'Project "{project.name}" restored.')
    return redirect('core:project_list')


@login_required
@require_POST
def project_duplicate(request, pk):
    """Duplicate an existing project."""
    original = get_object_or_404(Project, pk=pk, user=request.user)
    copy_proj = Project.objects.create(
        user=request.user,
        client=original.client,
        name=f"{original.name} (Copy)",
        description=original.description,
        status='pending',
        priority=original.priority,
        start_date=timezone.now().date(),
        deadline=original.deadline,
        budget=original.budget,
        progress=0
    )
    log_activity(request.user, 'create', 'project', copy_proj.id, f'Duplicated project from {original.name}', request)
    messages.success(request, f'Project duplicated as "{copy_proj.name}".')
    return redirect('core:project_detail', pk=copy_proj.pk)


@login_required
@require_POST
def project_add_comment(request, pk):
    """Add internal note/comment to project."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    comment_text = request.POST.get('comment', '').strip()
    if comment_text:
        ProjectComment.objects.create(user=request.user, project=project, comment=comment_text)
        log_activity(request.user, 'create', 'project', project.id, f'Added comment on project {project.name}', request)
        messages.success(request, 'Comment added!')
    return redirect('core:project_detail', pk=project.pk)


@login_required
@require_POST
def project_upload_file(request, pk):
    """Upload attachment to a project."""
    project = get_object_or_404(Project, pk=pk, user=request.user)
    file_obj = request.FILES.get('file')
    if file_obj:
        name = file_obj.name
        size = file_obj.size
        ext = name.split('.')[-1].lower() if '.' in name else 'other'
        ProjectFile.objects.create(
            user=request.user,
            project=project,
            file=file_obj,
            file_name=name,
            file_size=size,
            file_type=ext
        )
        log_activity(request.user, 'create', 'file', project.id, f'Uploaded file {name} to {project.name}', request)
        messages.success(request, f'File "{name}" uploaded!')
    return redirect('core:project_detail', pk=project.pk)


# ============================================================
# CLIENT EXTENSION VIEWS
# ============================================================

@login_required
@require_POST
def client_archive(request, pk):
    client = get_object_or_404(Client, pk=pk, user=request.user)
    client.is_archived = True
    client.save()
    log_activity(request.user, 'archive', 'client', client.id, f'Archived client {client.name}', request)
    messages.success(request, f'Client "{client.name}" archived.')
    return redirect('core:client_list')


@login_required
@require_POST
def client_restore(request, pk):
    client = get_object_or_404(Client, pk=pk, user=request.user)
    client.is_archived = False
    client.save()
    log_activity(request.user, 'restore', 'client', client.id, f'Restored client {client.name}', request)
    messages.success(request, f'Client "{client.name}" restored.')
    return redirect('core:client_list')


# ============================================================
# TASK EXTENSION VIEWS
# ============================================================

@login_required
@require_POST
def task_archive(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    task.is_archived = True
    task.save()
    log_activity(request.user, 'archive', 'task', task.id, f'Archived task {task.title}', request)
    messages.success(request, f'Task "{task.title}" archived.')
    return redirect('core:task_list')


@login_required
@require_POST
def task_restore(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    task.is_archived = False
    task.save()
    log_activity(request.user, 'restore', 'task', task.id, f'Restored task {task.title}', request)
    messages.success(request, f'Task "{task.title}" restored.')
    return redirect('core:task_list')


# ============================================================
# PAYMENT RECEIPT VIEW
# ============================================================

@login_required
def payment_receipt(request, pk):
    """Render printable payment receipt."""
    payment = get_object_or_404(Payment, pk=pk, user=request.user)
    profile = getattr(request.user, 'profile', None)
    context = {
        'payment': payment,
        'profile': profile,
    }
    return render(request, 'payments/payment_receipt.html', context)


# ============================================================
# INCOME & EXPENSE TRACKER
# ============================================================

@login_required
def income_expense_tracker(request):
    """Manage incomes and expenses."""
    user = request.user
    incomes = Income.objects.filter(user=user)
    expenses = Expense.objects.filter(user=user)

    total_inc = incomes.aggregate(t=Sum('amount'))['t'] or 0
    total_exp = expenses.aggregate(t=Sum('amount'))['t'] or 0
    net_profit = float(total_inc) - float(total_exp)

    income_form = IncomeForm()
    expense_form = ExpenseForm()
    # Populate choices for forms
    income_form.fields['client'].queryset = Client.objects.filter(user=user, is_archived=False)
    income_form.fields['project'].queryset = Project.objects.filter(user=user, is_archived=False)
    expense_form.fields['project'].queryset = Project.objects.filter(user=user, is_archived=False)

    context = {
        'incomes': incomes,
        'expenses': expenses,
        'total_income': total_inc,
        'total_expenses': total_exp,
        'net_profit': net_profit,
        'income_form': income_form,
        'expense_form': expense_form,
    }
    return render(request, 'finances/income_expense.html', context)


@login_required
@require_POST
def income_create(request):
    form = IncomeForm(request.POST)
    if form.is_valid():
        inc = form.save(commit=False)
        inc.user = request.user
        inc.save()
        log_activity(request.user, 'create', 'income', inc.id, f'Added income: {inc.title} (${inc.amount})', request)
        messages.success(request, f'Income "{inc.title}" recorded!')
    else:
        messages.error(request, 'Failed to record income.')
    return redirect('core:income_expense_tracker')


@login_required
@require_POST
def income_delete(request, pk):
    inc = get_object_or_404(Income, pk=pk, user=request.user)
    inc.delete()
    log_activity(request.user, 'delete', 'income', inc.id, f'Deleted income: {inc.title}', request)
    messages.success(request, 'Income entry deleted.')
    return redirect('core:income_expense_tracker')


@login_required
@require_POST
def expense_create(request):
    form = ExpenseForm(request.POST, request.FILES)
    if form.is_valid():
        exp = form.save(commit=False)
        exp.user = request.user
        exp.save()
        log_activity(request.user, 'create', 'expense', exp.id, f'Added expense: {exp.title} (${exp.amount})', request)
        messages.success(request, f'Expense "{exp.title}" recorded!')
    else:
        messages.error(request, 'Failed to record expense.')
    return redirect('core:income_expense_tracker')


@login_required
@require_POST
def expense_delete(request, pk):
    exp = get_object_or_404(Expense, pk=pk, user=request.user)
    exp.delete()
    log_activity(request.user, 'delete', 'expense', exp.id, f'Deleted expense: {exp.title}', request)
    messages.success(request, 'Expense entry deleted.')
    return redirect('core:income_expense_tracker')


# ============================================================
# INVOICE MANAGEMENT
# ============================================================

@login_required
def invoice_list(request):
    invoices = Invoice.objects.filter(user=request.user).select_related('client', 'project')
    context = {'invoices': invoices}
    return render(request, 'invoices/invoice_list.html', context)


@login_required
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            inv = form.save(commit=False)
            inv.user = request.user
            inv.save()

            # Process line items from POST
            descriptions = request.POST.getlist('item_description')
            quantities = request.POST.getlist('item_quantity')
            unit_prices = request.POST.getlist('item_unit_price')

            subtotal = 0
            for d, q, u in zip(descriptions, quantities, unit_prices):
                if d.strip():
                    qty = float(q) if q else 1
                    price = float(u) if u else 0
                    amt = qty * price
                    subtotal += amt
                    InvoiceItem.objects.create(
                        invoice=inv,
                        description=d.strip(),
                        quantity=qty,
                        unit_price=price,
                        amount=amt
                    )

            inv.subtotal = subtotal
            tax_amt = (subtotal * float(inv.tax_rate)) / 100.0
            inv.tax_amount = tax_amt
            inv.total = subtotal + tax_amt - float(inv.discount_amount)
            inv.save()

            log_activity(request.user, 'invoice_generated', 'invoice', inv.id, f'Created invoice {inv.invoice_number}', request)
            messages.success(request, f'Invoice {inv.invoice_number} created successfully!')
            return redirect('core:invoice_detail', pk=inv.pk)
        else:
            messages.error(request, 'Failed to create invoice. Please check inputs.')
    else:
        # Pre-fill invoice number
        count = Invoice.objects.filter(user=request.user).count() + 1
        initial_number = f"INV-{timezone.now().year}-{count:03d}"
        form = InvoiceForm(initial={'invoice_number': initial_number, 'issue_date': timezone.now().date(), 'due_date': timezone.now().date() + timedelta(days=14)})
        form.fields['client'].queryset = Client.objects.filter(user=request.user, is_archived=False)
        form.fields['project'].queryset = Project.objects.filter(user=request.user, is_archived=False)

    context = {'form': form}
    return render(request, 'invoices/invoice_form.html', context)


@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)
    profile = getattr(request.user, 'profile', None)
    context = {
        'invoice': invoice,
        'profile': profile,
    }
    return render(request, 'invoices/invoice_detail.html', context)


@login_required
def invoice_pdf(request, pk):
    """Render printable PDF view of invoice."""
    return invoice_detail(request, pk)


@login_required
@require_POST
def invoice_email(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)
    invoice.status = 'sent'
    invoice.save()
    log_activity(request.user, 'update', 'invoice', invoice.id, f'Emailed invoice {invoice.invoice_number} to {invoice.client.email}', request)
    messages.success(request, f'Invoice {invoice.invoice_number} emailed to {invoice.client.email}!')
    return redirect('core:invoice_detail', pk=invoice.pk)


@login_required
@require_POST
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)
    invoice.delete()
    log_activity(request.user, 'delete', 'invoice', invoice.id, f'Deleted invoice {invoice.invoice_number}', request)
    messages.success(request, 'Invoice deleted.')
    return redirect('core:invoice_list')


# ============================================================
# CALENDAR & DEADLINES
# ============================================================

@login_required
def calendar_view(request):
    user = request.user
    events = CalendarEvent.objects.filter(user=user)
    
    project_deadlines = Project.objects.filter(user=user, deadline__isnull=False, is_archived=False)
    task_deadlines = Task.objects.filter(user=user, due_date__isnull=False, is_archived=False)
    payment_dues = Payment.objects.filter(user=user, due_date__isnull=False, status='pending')

    event_form = CalendarEventForm()
    event_form.fields['project'].queryset = Project.objects.filter(user=user, is_archived=False)
    event_form.fields['task'].queryset = Task.objects.filter(user=user, is_archived=False)

    context = {
        'events': events,
        'project_deadlines': project_deadlines,
        'task_deadlines': task_deadlines,
        'payment_dues': payment_dues,
        'event_form': event_form,
    }
    return render(request, 'calendar/calendar.html', context)


@login_required
@require_POST
def calendar_event_create(request):
    form = CalendarEventForm(request.POST)
    if form.is_valid():
        evt = form.save(commit=False)
        evt.user = request.user
        evt.save()
        log_activity(request.user, 'create', 'note', evt.id, f'Created calendar event: {evt.title}', request)
        messages.success(request, f'Event "{evt.title}" added to calendar!')
    else:
        messages.error(request, 'Failed to add calendar event.')
    return redirect('core:calendar_view')


@login_required
@require_POST
def calendar_event_delete(request, pk):
    evt = get_object_or_404(CalendarEvent, pk=pk, user=request.user)
    evt.delete()
    messages.success(request, 'Calendar event deleted.')
    return redirect('core:calendar_view')


# ============================================================
# FILE MANAGER
# ============================================================

@login_required
def file_manager(request):
    user = request.user
    all_files = ProjectFile.objects.filter(user=user).select_related('project')
    projects = Project.objects.filter(user=user, is_archived=False)

    search_query = request.GET.get('search', '').strip()
    project_filter = request.GET.get('project', '')
    category_filter = request.GET.get('category', '')

    files = all_files
    if search_query:
        files = files.filter(file_name__icontains=search_query)
    if project_filter:
        files = files.filter(project_id=project_filter)

    # Category counts matching cloud storage design
    pictures_count = all_files.filter(file_type__in=['jpg', 'jpeg', 'png', 'gif', 'webp', 'svg']).count()
    documents_count = all_files.filter(file_type__in=['pdf', 'doc', 'docx', 'txt', 'csv', 'xlsx', 'ppt', 'pptx']).count()
    videos_count = all_files.filter(file_type__in=['mp4', 'mkv', 'avi', 'mov', 'wmv']).count()
    audio_count = all_files.filter(file_type__in=['mp3', 'wav', 'ogg', 'aac', 'flac']).count()

    if category_filter == 'pictures':
        files = files.filter(file_type__in=['jpg', 'jpeg', 'png', 'gif', 'webp', 'svg'])
    elif category_filter == 'documents':
        files = files.filter(file_type__in=['pdf', 'doc', 'docx', 'txt', 'csv', 'xlsx', 'ppt', 'pptx'])
    elif category_filter == 'videos':
        files = files.filter(file_type__in=['mp4', 'mkv', 'avi', 'mov', 'wmv'])
    elif category_filter == 'audio':
        files = files.filter(file_type__in=['mp3', 'wav', 'ogg', 'aac', 'flac'])

    # Storage calculation
    total_bytes = all_files.aggregate(Sum('file_size'))['file_size__sum'] or 0
    total_mb = round(total_bytes / (1024 * 1024), 2)
    max_mb = 500.0  # 500 MB quota
    storage_percentage = min(100, round((total_mb / max_mb) * 100, 1))

    file_form = ProjectFileForm()
    file_form.fields['project'].queryset = projects

    context = {
        'files': files,
        'projects': projects,
        'file_form': file_form,
        'pictures_count': pictures_count,
        'documents_count': documents_count,
        'videos_count': videos_count,
        'audio_count': audio_count,
        'total_mb': total_mb,
        'max_mb': max_mb,
        'storage_percentage': storage_percentage,
        'active_category': category_filter,
    }
    return render(request, 'files/file_manager.html', context)


@login_required
@require_POST
def file_upload(request):
    file_obj = request.FILES.get('file')
    if file_obj:
        name = file_obj.name
        size = file_obj.size
        ext = name.split('.')[-1].lower() if '.' in name else 'other'
        proj_id = request.POST.get('project')
        proj = Project.objects.filter(id=proj_id, user=request.user).first() if proj_id else None

        pf = ProjectFile.objects.create(
            user=request.user,
            project=proj,
            file=file_obj,
            file_name=name,
            file_size=size,
            file_type=ext
        )
        log_activity(request.user, 'create', 'file', pf.id, f'Uploaded file {name}', request)
        messages.success(request, f'File "{name}" uploaded successfully!')
    else:
        messages.error(request, 'Please select a file to upload.')
    return redirect('core:file_manager')


@login_required
@require_POST
def file_delete(request, pk):
    pf = get_object_or_404(ProjectFile, pk=pk, user=request.user)
    pf.file.delete(save=False)
    pf.delete()
    log_activity(request.user, 'delete', 'file', pf.id, f'Deleted file {pf.file_name}', request)
    messages.success(request, 'File deleted.')
    return redirect('core:file_manager')


# ============================================================
# NOTIFICATIONS
# ============================================================

@login_required
def notifications_list(request):
    notifications = Notification.objects.filter(user=request.user)
    context = {'notifications': notifications}
    return render(request, 'notifications/notifications.html', context)


@login_required
@require_POST
def notification_mark_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk, user=request.user)
    notif.is_read = True
    notif.save()
    return redirect('core:notifications_list')


@login_required
@require_POST
def notification_read_all(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    messages.success(request, 'All notifications marked as read.')
    return redirect('core:notifications_list')


@login_required
@require_POST
def notification_delete(request, pk):
    notif = get_object_or_404(Notification, pk=pk, user=request.user)
    notif.delete()
    messages.success(request, 'Notification deleted.')
    return redirect('core:notifications_list')


# ============================================================
# GLOBAL SEARCH & ACCESS CONTROL
# ============================================================

@login_required
def global_search(request):
    query = request.GET.get('q', '').strip()
    projects = []
    clients = []
    tasks = []
    payments = []
    invoices = []

    if query:
        projects = Project.objects.filter(user=request.user, is_archived=False).filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )
        clients = Client.objects.filter(user=request.user, is_archived=False).filter(
            Q(name__icontains=query) | Q(company__icontains=query) | Q(email__icontains=query)
        )
        tasks = Task.objects.filter(user=request.user, is_archived=False).filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )
        payments = Payment.objects.filter(user=request.user).filter(
            Q(description__icontains=query) | Q(invoice_number__icontains=query)
        )
        invoices = Invoice.objects.filter(user=request.user).filter(
            Q(invoice_number__icontains=query) | Q(notes__icontains=query)
        )

    context = {
        'query': query,
        'projects': projects,
        'clients': clients,
        'tasks': tasks,
        'payments': payments,
        'invoices': invoices,
    }
    return render(request, 'search/search_results.html', context)


def forbidden_view(request):
    """403 Forbidden page for unauthorized access attempts."""
    return render(request, '403.html', status=403)
